"""
Sincronizador de Precio y Costo: Pret a Home → Casa Lavan
CON CHECKPOINT — retoma desde donde quedó si se interrumpe

Uso:
    python3 sincronizar_precios_lavan.py             # Modo simulación
    python3 sincronizar_precios_lavan.py --ejecutar  # Aplica cambios reales
    python3 sincronizar_precios_lavan.py --reset     # Borra checkpoint y empieza de cero
"""

import requests, time, sys, json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DRY_RUN   = "--ejecutar" not in sys.argv
DO_RESET  = "--reset" in sys.argv
CHECKPOINT_FILE = "sync_checkpoint.json"

STORES = {
    "pret":  {"store_id": "2625285", "token": "7bf4cde46764d96772079d8cb1d10cd644aa35a0", "name": "Pret a Home"},
    "lavan": {"store_id": "6709240", "token": "e01b9ebab49d1e00973700eb3e6a1985b016e4e7", "name": "Casa Lavan"},
}


# ── Checkpoint ────────────────────────────────────────────────────────────────

def load_checkpoint():
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "r") as f:
            return json.load(f)
    return None

def save_checkpoint(data):
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump(data, f)

def delete_checkpoint():
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)
        print("  Checkpoint borrado.")


# ── API ───────────────────────────────────────────────────────────────────────

def api_headers(token):
    return {
        "Authentication": f"bearer {token}",
        "User-Agent": "PriceSync (soporte@pretahome.com)",
        "Content-Type": "application/json",
    }

def get_all_products(store_id, token, store_name):
    products = []
    page = 1
    print(f"\nTrayendo todos los productos de {store_name}...")
    while True:
        url = f"https://api.tiendanube.com/v1/{store_id}/products"
        r = requests.get(url, headers=api_headers(token), params={"per_page": 200, "page": page})
        if r.status_code == 429:
            print("  Rate limit, esperando 5s...")
            time.sleep(5)
            continue
        if r.status_code != 200:
            print(f"  Error {r.status_code}: {r.text[:200]}")
            break
        data = r.json()
        if not data:
            break
        products.extend(data)
        print(f"  Pagina {page}: {len(data)} productos ({len(products)} total)")
        if len(data) < 200:
            break
        page += 1
        time.sleep(0.3)
    print(f"  Total: {len(products)} productos")
    return products

def get_name(raw):
    if isinstance(raw, dict):
        return raw.get("es") or next(iter(raw.values()), "") or ""
    return raw or ""

def to_float(val):
    try:
        return float(val) if val is not None else None
    except (ValueError, TypeError):
        return None

def build_pret_sku_map(pret_products):
    mapping = {}
    for p in pret_products:
        name = get_name(p.get("name", ""))
        for v in p.get("variants", []):
            sku = (v.get("sku") or "").strip()
            if sku:
                mapping[sku] = {
                    "product_name": name,
                    "price": to_float(v.get("price")),
                    "cost": to_float(v.get("cost")),
                }
    return mapping

def update_variant(store_id, token, product_id, variant_id, payload):
    url = f"https://api.tiendanube.com/v1/{store_id}/products/{product_id}/variants/{variant_id}"
    r = requests.put(url, headers=api_headers(token), json=payload)
    if r.status_code == 429:
        print("  Rate limit en update, esperando 10s...")
        time.sleep(10)
        r = requests.put(url, headers=api_headers(token), json=payload)
    return r.status_code, r.json() if r.content else {}


# ── Sincronización con checkpoint ─────────────────────────────────────────────

def sincronizar(lavan_products, pret_map, dry_run=True):
    store_id = STORES["lavan"]["store_id"]
    token    = STORES["lavan"]["token"]

    # Cargar checkpoint si existe
    cp = load_checkpoint() if not dry_run else None
    done_variant_ids = set(cp["done_variant_ids"]) if cp else set()
    actualizados = cp["actualizados"] if cp else []
    sin_cambio   = cp["sin_cambio"]   if cp else []
    no_cruzados  = cp["no_cruzados"]  if cp else []
    errores      = cp["errores"]      if cp else []

    if cp:
        print(f"\nRetomando desde checkpoint — {len(done_variant_ids)} variantes ya procesadas")

    total_variantes = sum(len(p.get("variants", [])) for p in lavan_products)
    procesadas = len(done_variant_ids)

    print(f"\n{'SIMULACION' if dry_run else 'EJECUTANDO'} — {total_variantes} variantes totales | {procesadas} ya hechas\n")

    for p in lavan_products:
        product_name = get_name(p.get("name", ""))
        product_id   = p.get("id")
        published    = p.get("published", False)

        for v in p.get("variants", []):
            variant_id = v.get("id")
            procesadas += 1

            # Saltear si ya estaba en checkpoint
            if variant_id in done_variant_ids:
                continue

            sku        = (v.get("sku") or "").strip()
            lavan_price = to_float(v.get("price"))
            lavan_cost  = to_float(v.get("cost"))

            parts = []
            for attr in v.get("values", []):
                val = get_name(attr) if isinstance(attr, dict) else str(attr)
                if val:
                    parts.append(val)
            variant_desc = " / ".join(parts) if parts else "Sin variante"

            if not sku:
                no_cruzados.append({"sku": "", "product_name": product_name, "variant": variant_desc, "motivo": "Sin SKU", "published": published})
                done_variant_ids.add(variant_id)
                continue

            pret_info = pret_map.get(sku)
            if not pret_info:
                no_cruzados.append({"sku": sku, "product_name": product_name, "variant": variant_desc, "motivo": "No existe en Pret a Home", "published": published})
                done_variant_ids.add(variant_id)
                continue

            pret_price = pret_info["price"]
            pret_cost  = pret_info["cost"]

            price_igual = (lavan_price == pret_price)
            cost_igual  = (lavan_cost == pret_cost or (pret_cost is None and lavan_cost is None))

            if price_igual and cost_igual:
                sin_cambio.append({"sku": sku, "product_name": product_name, "variant": variant_desc, "price": lavan_price, "cost": lavan_cost, "published": published})
                done_variant_ids.add(variant_id)
                continue

            payload = {}
            if not price_igual and pret_price is not None:
                payload["price"] = str(pret_price)
            if not cost_igual and pret_cost is not None:
                payload["cost"] = str(pret_cost)

            row = {
                "sku": sku,
                "product_name": product_name,
                "variant": variant_desc,
                "lavan_price_antes": lavan_price,
                "lavan_cost_antes": lavan_cost,
                "pret_price": pret_price,
                "pret_cost": pret_cost,
                "published": published,
            }

            if dry_run:
                row["resultado"] = "SIMULADO"
                actualizados.append(row)
            else:
                status, resp = update_variant(store_id, token, product_id, variant_id, payload)
                if status in (200, 201):
                    row["resultado"] = "OK"
                    actualizados.append(row)
                else:
                    row["resultado"] = f"ERROR {status}"
                    row["error_detail"] = str(resp)[:200]
                    errores.append(row)
                time.sleep(0.2)

            done_variant_ids.add(variant_id)

            # Guardar checkpoint cada 50 variantes procesadas
            if not dry_run and procesadas % 50 == 0:
                save_checkpoint({
                    "done_variant_ids": list(done_variant_ids),
                    "actualizados": actualizados,
                    "sin_cambio": sin_cambio,
                    "no_cruzados": no_cruzados,
                    "errores": errores,
                })
                print(f"  Procesadas: {procesadas}/{total_variantes} | Actualizadas: {len(actualizados)} | Errores: {len(errores)}")

    if not dry_run:
        delete_checkpoint()

    print(f"\nProceso completado:")
    print(f"  Actualizadas: {len(actualizados)}")
    print(f"  Ya estaban igual: {len(sin_cambio)}")
    print(f"  No cruzadas: {len(no_cruzados)}")
    print(f"  Errores: {len(errores)}")

    return actualizados, sin_cambio, no_cruzados, errores


# ── Reporte Excel ─────────────────────────────────────────────────────────────

def create_report(actualizados, sin_cambio, no_cruzados, errores, dry_run):
    wb = openpyxl.Workbook()
    C_DARK = "2C3E50"; C_GREEN = "D5F5E3"; C_RED = "FADBD8"
    C_YELLOW = "FEF9E7"; C_ALT = "EAECEE"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    MONEY = '"$"#,##0.00'

    def hdr(ws, row, col, text, bg, fg="FFFFFF"):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Arial", bold=True, color=fg, size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    def val(ws, row, col, value, bg=None, bold=False, center=True, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", size=9, bold=bold)
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        c.border = border
        if bg: c.fill = PatternFill("solid", start_color=bg)
        if fmt and value is not None: c.number_format = fmt

    # Hoja 1 — Actualizados
    ws1 = wb.active
    ws1.title = "Simulacion" if dry_run else "Actualizados"
    ws1.freeze_panes = "A3"
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"{'SIMULACION — ' if dry_run else ''}Variantes actualizadas en Casa Lavan | {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", start_color="922B21" if dry_run else "1E8449")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 26
    for i, (h, bg) in enumerate(zip(
        ["SKU", "Producto", "Variante", "Estado Lavan", "Precio Lavan ANTES", "Costo Lavan ANTES", "Precio Pret (nuevo)", "Costo Pret (nuevo)", "Resultado"],
        [C_DARK]*4 + ["1A5276"]*2 + ["1E8449"]*2 + [C_DARK]
    ), 1):
        hdr(ws1, 2, i, h, bg)
    ws1.row_dimensions[2].height = 28
    for ri, r in enumerate(actualizados, 3):
        alt = C_ALT if ri % 2 == 0 else None
        val(ws1, ri, 1, r["sku"], alt, center=False)
        val(ws1, ri, 2, r["product_name"], alt, center=False)
        val(ws1, ri, 3, r["variant"], alt, center=False)
        val(ws1, ri, 4, "Publico" if r["published"] else "Oculto", alt)
        val(ws1, ri, 5, r["lavan_price_antes"], alt, fmt=MONEY)
        val(ws1, ri, 6, r["lavan_cost_antes"], alt, fmt=MONEY)
        val(ws1, ri, 7, r["pret_price"], C_GREEN, fmt=MONEY)
        val(ws1, ri, 8, r["pret_cost"], C_GREEN, fmt=MONEY)
        val(ws1, ri, 9, r["resultado"], C_YELLOW if r["resultado"] == "SIMULADO" else C_GREEN, bold=True)
    for i, w in enumerate([14, 42, 25, 10, 18, 18, 18, 18, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Hoja 2 — Ya estaban igual
    ws2 = wb.create_sheet("Ya estaban igual")
    ws2.freeze_panes = "A3"
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"Ya tenian el mismo precio y costo | {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws2["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", start_color="1E8449")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26
    for i, h in enumerate(["SKU", "Producto", "Variante", "Estado Lavan", "Precio", "Costo"], 1):
        hdr(ws2, 2, i, h, C_DARK)
    for ri, r in enumerate(sin_cambio, 3):
        alt = C_ALT if ri % 2 == 0 else None
        val(ws2, ri, 1, r["sku"], alt, center=False)
        val(ws2, ri, 2, r["product_name"], alt, center=False)
        val(ws2, ri, 3, r["variant"], alt, center=False)
        val(ws2, ri, 4, "Publico" if r["published"] else "Oculto", alt)
        val(ws2, ri, 5, r["price"], C_GREEN, fmt=MONEY)
        val(ws2, ri, 6, r["cost"], C_GREEN, fmt=MONEY)
    for i, w in enumerate([14, 42, 25, 10, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Hoja 3 — No cruzados
    ws3 = wb.create_sheet("No cruzados")
    ws3.freeze_panes = "A3"
    ws3.merge_cells("A1:E1")
    ws3["A1"] = f"Sin cruce en Pret — no modificadas | {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws3["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", start_color=C_DARK)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26
    for i, h in enumerate(["SKU", "Producto", "Variante", "Estado Lavan", "Motivo"], 1):
        hdr(ws3, 2, i, h, C_DARK)
    for ri, r in enumerate(no_cruzados, 3):
        alt = C_ALT if ri % 2 == 0 else None
        val(ws3, ri, 1, r["sku"], alt, center=False)
        val(ws3, ri, 2, r["product_name"], alt, center=False)
        val(ws3, ri, 3, r["variant"], alt, center=False)
        val(ws3, ri, 4, "Publico" if r["published"] else "Oculto", alt)
        val(ws3, ri, 5, r["motivo"], C_YELLOW, center=False)
    for i, w in enumerate([14, 42, 25, 10, 30], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Hoja 4 — Errores
    if errores:
        ws_e = wb.create_sheet("Errores")
        ws_e.merge_cells("A1:E1")
        ws_e["A1"] = "Variantes con error"
        ws_e["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws_e["A1"].fill = PatternFill("solid", start_color="922B21")
        ws_e["A1"].alignment = Alignment(horizontal="center", vertical="center")
        for i, h in enumerate(["SKU", "Producto", "Variante", "Resultado", "Detalle"], 1):
            hdr(ws_e, 2, i, h, "922B21")
        for ri, r in enumerate(errores, 3):
            alt = C_ALT if ri % 2 == 0 else None
            val(ws_e, ri, 1, r["sku"], alt, center=False)
            val(ws_e, ri, 2, r["product_name"], alt, center=False)
            val(ws_e, ri, 3, r["variant"], alt, center=False)
            val(ws_e, ri, 4, r["resultado"], C_RED, bold=True)
            val(ws_e, ri, 5, r.get("error_detail", ""), C_RED, center=False)
        for i, w in enumerate([14, 42, 25, 12, 50], 1):
            ws_e.column_dimensions[get_column_letter(i)].width = w

    # Hoja 5 — Resumen
    ws5 = wb.create_sheet("Resumen")
    ws5.column_dimensions["A"].width = 50
    ws5.column_dimensions["B"].width = 14
    thin2 = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    def srow(row, label, value, bg=None):
        for col, v in [(1, label), (2, value)]:
            c = ws5.cell(row=row, column=col, value=v)
            c.font = Font(name="Arial", size=10, bold=(col == 1))
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            c.border = thin2
            if bg: c.fill = PatternFill("solid", start_color=bg)
        ws5.row_dimensions[row].height = 22

    ws5.merge_cells("A1:B1")
    ws5["A1"] = f"Resumen {'SIMULACION' if dry_run else 'EJECUCION'} — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws5["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws5["A1"].fill = PatternFill("solid", start_color=C_DARK)
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 28
    srow(2, "Variantes actualizadas" + (" (simulacion)" if dry_run else ""), len(actualizados), C_YELLOW if dry_run else C_GREEN)
    srow(3, "Ya tenian el mismo precio y costo", len(sin_cambio), C_GREEN)
    srow(4, "No cruzadas con Pret (sin cambio)", len(no_cruzados), "F2F3F4")
    srow(5, "Errores", len(errores), C_RED if errores else C_GREEN)

    suffix = "simulacion" if dry_run else "ejecutado"
    filename = f"sync_precios_lavan_{suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    return filename


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if DO_RESET:
        delete_checkpoint()
        print("Checkpoint borrado. Corré el script de nuevo para empezar desde cero.")
        sys.exit(0)

    if DRY_RUN:
        print("=" * 60)
        print("  MODO SIMULACION — no se cambia nada en Lavan")
        print("  Para aplicar: python3 sincronizar_precios_lavan.py --ejecutar")
        print("=" * 60)
    else:
        cp = load_checkpoint()
        if cp:
            print("=" * 60)
            print(f"  RETOMANDO desde checkpoint — {len(cp['done_variant_ids'])} variantes ya procesadas")
            print("  Para empezar de cero: python3 sincronizar_precios_lavan.py --reset")
            print("=" * 60)
        else:
            print("=" * 60)
            print("  MODO EJECUCION REAL — actualizando precios en Lavan")
            print("=" * 60)

    pret_products  = get_all_products(STORES["pret"]["store_id"],  STORES["pret"]["token"],  "Pret a Home")
    lavan_products = get_all_products(STORES["lavan"]["store_id"], STORES["lavan"]["token"], "Casa Lavan")

    print("\nConstruyendo mapa de precios de Pret...")
    pret_map = build_pret_sku_map(pret_products)
    print(f"  {len(pret_map)} SKUs en Pret a Home")

    actualizados, sin_cambio, no_cruzados, errores = sincronizar(lavan_products, pret_map, dry_run=DRY_RUN)

    print("\nGenerando reporte Excel...")
    filename = create_report(actualizados, sin_cambio, no_cruzados, errores, dry_run=DRY_RUN)
    print(f"  Reporte: {filename}")

    if DRY_RUN:
        print(f"\nSIMULACION completada. Para aplicar los cambios reales:")
        print(f"  python3 sincronizar_precios_lavan.py --ejecutar")
