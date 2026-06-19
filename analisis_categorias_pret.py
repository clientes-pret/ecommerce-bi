"""
Análisis de Categorías — Pret a Home (solo productos públicos)
- Lista todas las categorías con cantidad de productos activos
- Detecta productos con la misma categoría asignada dos veces
- Genera Excel para curación

Uso:
    python3 analisis_categorias_pret.py
"""

import requests
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

STORE_ID = "2625285"
TOKEN    = "7bf4cde46764d96772079d8cb1d10cd644aa35a0"

def api_headers():
    return {
        "Authentication": f"bearer {TOKEN}",
        "User-Agent": "CatAnalysis (soporte@pretahome.com)",
        "Content-Type": "application/json",
    }

def get_name(raw):
    if isinstance(raw, dict):
        return raw.get("es") or next(iter(raw.values()), "") or ""
    return raw or ""

def get_all_products():
    products = []
    page = 1
    print("\nTrayendo productos públicos de Pret a Home...")
    while True:
        url = f"https://api.tiendanube.com/v1/{STORE_ID}/products"
        r = requests.get(url, headers=api_headers(), params={"per_page": 200, "page": page, "published": "true"})
        if r.status_code == 429:
            time.sleep(5)
            continue
        if r.status_code != 200:
            print(f"  Error {r.status_code}: {r.text[:200]}")
            break
        data = r.json()
        if not data:
            break
        products.extend(data)
        print(f"  Página {page}: {len(data)} productos ({len(products)} total)")
        if len(data) < 200:
            break
        page += 1
        time.sleep(0.3)
    print(f"  Total: {len(products)} productos públicos")
    return products

def get_all_categories():
    """Trae todas las categorías con su nombre e ID."""
    cats = {}
    page = 1
    print("\nTrayendo categorías...")
    while True:
        url = f"https://api.tiendanube.com/v1/{STORE_ID}/categories"
        r = requests.get(url, headers=api_headers(), params={"per_page": 200, "page": page})
        if r.status_code == 429:
            time.sleep(5)
            continue
        if r.status_code != 200:
            print(f"  Error {r.status_code}: {r.text[:200]}")
            break
        data = r.json()
        if not data:
            break
        for c in data:
            cats[c["id"]] = {
                "name": get_name(c.get("name", "")),
                "id": c["id"],
                "parent": c["parent"]["id"] if isinstance(c.get("parent"), dict) else c.get("parent") if isinstance(c.get("parent"), int) else None,
            }
        print(f"  Página {page}: {len(data)} categorías")
        if len(data) < 200:
            break
        page += 1
        time.sleep(0.3)
    print(f"  Total: {len(cats)} categorías")
    return cats

def analyze(products, categories):
    # cat_id → lista de productos
    cat_products = defaultdict(list)
    # producto → lista de cat_ids asignadas
    product_cats = {}
    # productos con categoría duplicada
    duplicados = []

    for p in products:
        pid = p.get("id")
        pname = get_name(p.get("name", ""))
        sku = ""
        variants = p.get("variants", [])
        if variants:
            sku = (variants[0].get("sku") or "").strip()

        cat_ids = [c.get("id") for c in p.get("categories", []) if c.get("id")]
        product_cats[pid] = cat_ids

        # Detectar duplicados: mismo cat_id más de una vez
        seen = set()
        dup_cats = []
        for cid in cat_ids:
            if cid in seen:
                dup_cats.append(cid)
            seen.add(cid)

        if dup_cats:
            for cid in set(dup_cats):
                cname = categories.get(cid, {}).get("name", f"ID {cid}")
                duplicados.append({
                    "product_id": pid,
                    "product_name": pname,
                    "sku": sku,
                    "cat_id": cid,
                    "cat_name": cname,
                    "veces": cat_ids.count(cid),
                })

        # Acumular en cat_products
        for cid in set(cat_ids):  # set para no contar duplicados en el total
            cat_products[cid].append({
                "product_id": pid,
                "product_name": pname,
                "sku": sku,
                "n_cats": len(set(cat_ids)),
            })

    # Construir tabla de categorías
    cat_rows = []
    for cid, cat_info in categories.items():
        prods = cat_products.get(cid, [])
        parent_id = cat_info.get("parent")
        parent_name = categories.get(parent_id, {}).get("name", "") if parent_id else ""
        cat_rows.append({
            "cat_id": cid,
            "cat_name": cat_info["name"],
            "parent_id": parent_id or "",
            "parent_name": parent_name,
            "n_products": len(prods),
            "products": prods,
        })

    # Ordenar: primero por padre, luego por nombre
    cat_rows.sort(key=lambda r: (r["parent_name"] or r["cat_name"], r["cat_name"]))

    # Productos sin ninguna categoría
    sin_categoria = [
        p for p in products
        if not p.get("categories")
    ]

    return cat_rows, duplicados, sin_categoria


def create_excel(cat_rows, duplicados, sin_categoria, products):
    wb = openpyxl.Workbook()

    C_DARK   = "2C3E50"
    C_GREEN  = "D5F5E3"
    C_RED    = "FADBD8"
    C_YELLOW = "FEF9E7"
    C_BLUE   = "D6EAF8"
    C_ALT    = "F2F3F4"
    C_GREY   = "EAECEE"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, col, text, bg, fg="FFFFFF"):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Arial", bold=True, color=fg, size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    def val(ws, row, col, value, bg=None, bold=False, center=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", size=9, bold=bold)
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        c.border = border
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        return c

    def make_title(ws, text, bg, cols="A1:F1"):
        ws.merge_cells(cols)
        ws[cols.split(":")[0]] = text
        ws[cols.split(":")[0]].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws[cols.split(":")[0]].fill = PatternFill("solid", start_color=bg)
        ws[cols.split(":")[0]].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    # ── Hoja 1: Todas las categorías ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Categorias"
    ws1.freeze_panes = "A3"
    make_title(ws1, f"Categorías de Pret a Home — Solo productos públicos | {datetime.now().strftime('%d/%m/%Y %H:%M')}", C_DARK, "A1:F1")

    for i, (h, bg) in enumerate(zip(
        ["ID Categoría", "Categoría Padre", "Nombre Categoría", "Productos Activos", "% del Total", "Observación"],
        [C_DARK, C_DARK, C_DARK, C_DARK, C_DARK, C_DARK]
    ), 1):
        hdr(ws1, 2, i, h, bg)
    ws1.row_dimensions[2].height = 28

    total_products = len(products)
    for ri, r in enumerate(cat_rows, 3):
        alt = C_ALT if ri % 2 == 0 else None
        n = r["n_products"]
        pct = n / total_products * 100 if total_products else 0

        # Observación
        if n == 0:
            obs = "Categoría vacía — sin productos activos"
            obs_bg = C_YELLOW
        elif n == 1:
            obs = "Solo 1 producto"
            obs_bg = C_BLUE
        else:
            obs = ""
            obs_bg = alt

        val(ws1, ri, 1, r["cat_id"], alt)
        val(ws1, ri, 2, r["parent_name"], alt, center=False)
        val(ws1, ri, 3, r["cat_name"], alt, center=False, bold=not r["parent_name"])
        c4 = val(ws1, ri, 4, n, C_RED if n == 0 else (C_BLUE if n == 1 else C_GREEN), bold=True)
        val(ws1, ri, 5, f"{pct:.1f}%", alt)
        val(ws1, ri, 6, obs, obs_bg, center=False)

    for i, w in enumerate([12, 28, 35, 16, 10, 35], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 2: Detalle por categoría (productos dentro de cada una) ──────────
    ws2 = wb.create_sheet("Detalle por Categoria")
    ws2.freeze_panes = "A3"
    make_title(ws2, "Detalle — qué productos hay en cada categoría", C_DARK, "A1:E1")

    for i, h in enumerate(["Categoría Padre", "Categoría", "ID Producto", "Producto", "SKU"], 1):
        hdr(ws2, 2, i, h, C_DARK)
    ws2.row_dimensions[2].height = 28

    ri = 3
    for r in cat_rows:
        if not r["products"]:
            continue
        for prod in sorted(r["products"], key=lambda x: x["product_name"]):
            alt = C_ALT if ri % 2 == 0 else None
            val(ws2, ri, 1, r["parent_name"], alt, center=False)
            val(ws2, ri, 2, r["cat_name"], alt, center=False)
            val(ws2, ri, 3, prod["product_id"], alt)
            val(ws2, ri, 4, prod["product_name"], alt, center=False)
            val(ws2, ri, 5, prod["sku"], alt, center=False)
            ri += 1

    for i, w in enumerate([28, 35, 12, 48, 15], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 3: Categorías duplicadas ─────────────────────────────────────────
    ws3 = wb.create_sheet("Cat Duplicadas en Producto")
    ws3.freeze_panes = "A3"
    make_title(ws3,
        f"Productos con la misma categoría asignada más de una vez ({len(duplicados)} casos)",
        "922B21" if duplicados else "1E8449", "A1:F1")

    for i, h in enumerate(["ID Producto", "Producto", "SKU", "ID Categoría", "Categoría Duplicada", "Veces asignada"], 1):
        hdr(ws3, 2, i, h, "922B21" if duplicados else C_DARK)
    ws3.row_dimensions[2].height = 28

    for ri, r in enumerate(duplicados, 3):
        alt = C_ALT if ri % 2 == 0 else None
        val(ws3, ri, 1, r["product_id"], alt)
        val(ws3, ri, 2, r["product_name"], alt, center=False)
        val(ws3, ri, 3, r["sku"], alt, center=False)
        val(ws3, ri, 4, r["cat_id"], alt)
        val(ws3, ri, 5, r["cat_name"], alt, center=False)
        val(ws3, ri, 6, r["veces"], C_RED, bold=True)

    if not duplicados:
        ws3.cell(row=3, column=1).value = "Sin categorías duplicadas — todo OK"
        ws3.cell(row=3, column=1).font = Font(name="Arial", size=10, bold=True, color="1E8449")

    for i, w in enumerate([12, 48, 15, 12, 35, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 4: Productos sin categoría ───────────────────────────────────────
    ws4 = wb.create_sheet("Sin Categoria")
    ws4.freeze_panes = "A3"
    make_title(ws4,
        f"Productos públicos sin ninguna categoría asignada ({len(sin_categoria)} productos)",
        "922B21" if sin_categoria else "1E8449", "A1:D1")

    for i, h in enumerate(["ID Producto", "Producto", "SKU", "Precio"], 1):
        hdr(ws4, 2, i, h, "922B21" if sin_categoria else C_DARK)
    ws4.row_dimensions[2].height = 28

    for ri, p in enumerate(sin_categoria, 3):
        alt = C_ALT if ri % 2 == 0 else None
        name = get_name(p.get("name", ""))
        variants = p.get("variants", [])
        sku = (variants[0].get("sku") or "").strip() if variants else ""
        price = variants[0].get("price") if variants else ""
        val(ws4, ri, 1, p.get("id"), alt)
        val(ws4, ri, 2, name, alt, center=False)
        val(ws4, ri, 3, sku, alt, center=False)
        val(ws4, ri, 4, price, alt)

    if not sin_categoria:
        ws4.cell(row=3, column=1).value = "Todos los productos tienen al menos una categoría"
        ws4.cell(row=3, column=1).font = Font(name="Arial", size=10, bold=True, color="1E8449")

    for i, w in enumerate([12, 48, 15, 14], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 5: Resumen ───────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Resumen")
    ws5.column_dimensions["A"].width = 45
    ws5.column_dimensions["B"].width = 14
    thin2 = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))

    def srow(row, label, value, bg=None):
        for col, v in [(1, label), (2, value)]:
            c = ws5.cell(row=row, column=col, value=v)
            c.font = Font(name="Arial", size=10, bold=(col == 1))
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            c.border = thin2
            if bg:
                c.fill = PatternFill("solid", start_color=bg)
        ws5.row_dimensions[row].height = 22

    ws5.merge_cells("A1:B1")
    ws5["A1"] = f"Resumen Categorías Pret a Home — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws5["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws5["A1"].fill = PatternFill("solid", start_color=C_DARK)
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 28

    n_cats_vacias = sum(1 for r in cat_rows if r["n_products"] == 0)
    n_cats_1prod  = sum(1 for r in cat_rows if r["n_products"] == 1)

    srow(2,  "Total productos públicos",                       total_products)
    srow(3,  "Total categorías",                               len(cat_rows))
    srow(4,  "Categorías con productos activos",               len(cat_rows) - n_cats_vacias, C_GREEN)
    srow(5,  "Categorías vacías (sin productos activos)",      n_cats_vacias,  C_YELLOW if n_cats_vacias else C_GREEN)
    srow(6,  "Categorías con solo 1 producto",                 n_cats_1prod,   C_BLUE)
    srow(7,  "Productos con categoría duplicada",              len(duplicados), C_RED if duplicados else C_GREEN)
    srow(8,  "Productos sin ninguna categoría",                len(sin_categoria), C_RED if sin_categoria else C_GREEN)

    filename = f"categorias_pret_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    return filename


if __name__ == "__main__":
    products   = get_all_products()
    categories = get_all_categories()

    print("\nAnalizando...")
    cat_rows, duplicados, sin_categoria = analyze(products, categories)

    n_vacias = sum(1 for r in cat_rows if r["n_products"] == 0)
    print(f"  Total categorías:          {len(cat_rows)}")
    print(f"  Categorías con productos:  {len(cat_rows) - n_vacias}")
    print(f"  Categorías vacías:         {n_vacias}")
    print(f"  Categorías duplicadas:     {len(duplicados)}")
    print(f"  Productos sin categoría:   {len(sin_categoria)}")

    print("\nGenerando Excel...")
    filename = create_excel(cat_rows, duplicados, sin_categoria, products)
    print(f"  Listo: {filename}")
    print(f"\n  sheets.google.com → Nuevo → Importar → '{filename}'")
