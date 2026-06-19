#!/usr/bin/env python3
"""
Categorías vendidas en Pret a Home (Tienda Nube)
Período: 8 de mayo 2026 → hoy
Genera un Google Sheet con análisis de exposición por categoría
"""

import requests
import time
import json
import sys
from datetime import datetime, date
from collections import defaultdict

# ── Credenciales Tienda Nube ─────────────────────────────────────────────────
STORE_ID    = "2625285"
TOKEN       = "7bf4cde46764d96772079d8cb1d10cd644aa35a0"
BASE_URL    = f"https://api.tiendanube.com/v1/{STORE_ID}"
HEADERS_TN  = {
    "Authentication": f"bearer {TOKEN}",
    "User-Agent": "PretAHome Analytics (analytics@pretahome.com)",
    "Content-Type": "application/json"
}

# ── Credenciales Google ──────────────────────────────────────────────────────
# Opción A: path al JSON de service account
# Opción B: dejá vacío y el script crea un Excel local como fallback
GOOGLE_SA_JSON = ""  # ← pegá acá el path, ej: "/Users/matiaslerer/Downloads/pret-a-home-xxxx.json"

DATE_FROM = "2026-05-08T00:00:00-03:00"

# ─────────────────────────────────────────────────────────────────────────────

def fetch_orders():
    """Descarga todas las órdenes pagas desde DATE_FROM"""
    all_orders = []
    page = 1
    params = {
        "per_page": 200,
        "created_at_min": DATE_FROM,
        "payment_status": "paid",
        "fields": "id,number,created_at,products,total,payment_status,status"
    }
    print(f"🔍 Buscando órdenes desde {DATE_FROM[:10]} ...")
    while True:
        params["page"] = page
        r = requests.get(f"{BASE_URL}/orders", headers=HEADERS_TN, params=params)
        if r.status_code != 200:
            print(f"❌ Error {r.status_code}: {r.text[:200]}")
            sys.exit(1)
        data = r.json()
        if not data:
            break
        valid = [o for o in data if o.get("status") != "cancelled"]
        all_orders.extend(valid)
        print(f"   Página {page}: {len(data)} órdenes ({len(valid)} válidas) — acum: {len(all_orders)}")
        if len(data) < 200:
            break
        page += 1
        time.sleep(0.5)
    print(f"✅ Total órdenes válidas: {len(all_orders)}\n")
    return all_orders


def fetch_product_categories():
    """Descarga todos los productos con sus categorías"""
    print("📦 Descargando catálogo de productos y categorías ...")
    products = {}
    page = 1
    params = {
        "per_page": 200,
        "fields": "id,name,categories,variants"
    }
    while True:
        params["page"] = page
        r = requests.get(f"{BASE_URL}/products", headers=HEADERS_TN, params=params)
        if r.status_code != 200:
            print(f"   ⚠️  Error al traer productos: {r.status_code}")
            break
        data = r.json()
        if not data:
            break
        for p in data:
            cats = [c.get("name", {}) for c in p.get("categories", [])]
            # name puede ser dict con idioma o string
            cat_names = []
            for c in p.get("categories", []):
                n = c.get("name", "")
                if isinstance(n, dict):
                    n = n.get("es", "") or next(iter(n.values()), "")
                cat_names.append(n)
            products[str(p["id"])] = {
                "name": p.get("name", {}).get("es", "") if isinstance(p.get("name"), dict) else p.get("name", ""),
                "categories": cat_names if cat_names else ["Sin categoría"]
            }
        print(f"   Página {page}: {len(data)} productos — acum: {len(products)}")
        if len(data) < 200:
            break
        page += 1
        time.sleep(0.5)
    print(f"✅ Catálogo cargado: {len(products)} productos\n")
    return products


def analyze_by_category(orders, products):
    """Agrupa las ventas por categoría"""
    cat_stats = defaultdict(lambda: {
        "ordenes": set(),
        "unidades": 0,
        "revenue": 0.0,
        "productos_set": set()
    })

    for order in orders:
        order_id = order.get("id")
        for item in order.get("products", []):
            prod_id = str(item.get("product_id", ""))
            qty = item.get("quantity", 1)
            price = float(item.get("price", 0))
            
            prod_info = products.get(prod_id, {
                "name": item.get("name", "Desconocido"),
                "categories": ["Sin categoría"]
            })
            
            for cat in prod_info["categories"]:
                cat_stats[cat]["ordenes"].add(order_id)
                cat_stats[cat]["unidades"] += qty
                cat_stats[cat]["revenue"] += price * qty
                cat_stats[cat]["productos_set"].add(prod_id)

    # Convertir sets a conteos
    result = []
    for cat, stats in cat_stats.items():
        result.append({
            "categoria": cat,
            "ordenes": len(stats["ordenes"]),
            "unidades_vendidas": stats["unidades"],
            "revenue_ars": round(stats["revenue"], 2),
            "productos_distintos": len(stats["productos_set"]),
            "ticket_promedio_unidad": round(stats["revenue"] / stats["unidades"], 2) if stats["unidades"] > 0 else 0
        })

    result.sort(key=lambda x: x["revenue_ars"], reverse=True)
    return result


def build_google_sheet(rows, sa_json_path, total_orders, date_from):
    """Crea el Google Sheet con los datos"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("📦 Instalando gspread y google-auth ...")
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "gspread", "google-auth", "--break-system-packages", "-q"])
        import gspread
        from google.oauth2.service_account import Credentials

    SCOPES = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    creds = Credentials.from_service_account_file(sa_json_path, scopes=SCOPES)
    gc = gspread.authorize(creds)

    today_str = date.today().strftime("%d/%m/%Y")
    sheet_title = f"Categorías Pret a Home — {date_from[:10]} al {today_str}"
    sh = gc.create(sheet_title)
    sh.share(None, perm_type="anyone", role="reader")

    ws = sh.sheet1
    ws.update_title("Por Categoría")

    # ── Encabezado informativo ────────────────────────────────────────────────
    ws.update("A1", [[f"Pret a Home (Tienda Nube) — Ventas por Categoría"]])
    ws.update("A2", [[f"Período: {date_from[:10]} al {today_str}   |   Total órdenes analizadas: {total_orders}"]])

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = [
        "Categoría",
        "Órdenes",
        "Unidades vendidas",
        "Revenue ARS",
        "Productos distintos",
        "Ticket prom. por unidad (ARS)",
        "% Revenue",
        "% Órdenes"
    ]
    ws.update("A4", [headers])

    # ── Datos ─────────────────────────────────────────────────────────────────
    total_rev  = sum(r["revenue_ars"] for r in rows)
    total_ord  = sum(r["ordenes"] for r in rows)

    data_rows = []
    for r in rows:
        pct_rev = round(r["revenue_ars"] / total_rev * 100, 1) if total_rev else 0
        pct_ord = round(r["ordenes"]     / total_ord * 100, 1) if total_ord else 0
        data_rows.append([
            r["categoria"],
            r["ordenes"],
            r["unidades_vendidas"],
            r["revenue_ars"],
            r["productos_distintos"],
            r["ticket_promedio_unidad"],
            pct_rev,
            pct_ord
        ])

    if data_rows:
        ws.update(f"A5:H{4 + len(data_rows)}", data_rows)

    # ── Fila de totales ───────────────────────────────────────────────────────
    tot_row = 5 + len(data_rows)
    ws.update(f"A{tot_row}", [["TOTAL", total_ord, sum(r["unidades_vendidas"] for r in rows), total_rev, "", "", 100.0, 100.0]])

    # ── Formato ───────────────────────────────────────────────────────────────
    import gspread.utils

    # Header row bold + color fondo oscuro
    ws.format("A4:H4", {
        "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        "backgroundColor": {"red": 0.18, "green": 0.24, "blue": 0.31},
        "horizontalAlignment": "CENTER"
    })
    # Título
    ws.format("A1", {"textFormat": {"bold": True, "fontSize": 14}})
    ws.format("A2", {"textFormat": {"italic": True, "foregroundColor": {"red": 0.4, "green": 0.4, "blue": 0.4}}})
    # Datos alternados
    for i, _ in enumerate(data_rows):
        row_num = 5 + i
        bg = {"red": 0.94, "green": 0.96, "blue": 0.98} if i % 2 == 0 else {"red": 1, "green": 1, "blue": 1}
        ws.format(f"A{row_num}:H{row_num}", {"backgroundColor": bg})
    # Fila total
    ws.format(f"A{tot_row}:H{tot_row}", {
        "textFormat": {"bold": True},
        "backgroundColor": {"red": 0.95, "green": 0.95, "blue": 0.85}
    })
    # Revenue como moneda
    ws.format(f"D5:D{tot_row}", {"numberFormat": {"type": "NUMBER", "pattern": "#,##0.00"}})
    ws.format(f"F5:F{tot_row}", {"numberFormat": {"type": "NUMBER", "pattern": "#,##0.00"}})
    # % con símbolo
    ws.format(f"G5:H{tot_row}", {"numberFormat": {"type": "NUMBER", "pattern": "0.0\"%\""}})

    # Anchos de columna
    from gspread_formatting import set_column_width
    try:
        set_column_width(ws, "A", 220)
        set_column_width(ws, "B", 90)
        set_column_width(ws, "C", 130)
        set_column_width(ws, "D", 140)
        set_column_width(ws, "E", 130)
        set_column_width(ws, "F", 180)
        set_column_width(ws, "G", 100)
        set_column_width(ws, "H", 100)
    except Exception:
        pass  # gspread_formatting opcional

    url = f"https://docs.google.com/spreadsheets/d/{sh.id}"
    return url


def build_excel_fallback(rows, total_orders, date_from):
    """Crea un Excel local si no hay credenciales de Google"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Por Categoría"

    today_str = date.today().strftime("%d/%m/%Y")

    # Título
    ws["A1"] = "Pret a Home (Tienda Nube) — Ventas por Categoría"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Período: {date_from[:10]} al {today_str}   |   Órdenes: {total_orders}"
    ws["A2"].font = Font(italic=True, color="666666")

    headers = ["Categoría", "Órdenes", "Unidades vendidas", "Revenue ARS",
               "Productos distintos", "Ticket prom. unidad (ARS)", "% Revenue", "% Órdenes"]
    
    header_fill = PatternFill("solid", fgColor="2E3E4F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_rev = sum(r["revenue_ars"] for r in rows)
    total_ord = sum(r["ordenes"] for r in rows)

    fills = [PatternFill("solid", fgColor="EFF4FA"), PatternFill("solid", fgColor="FFFFFF")]

    for i, r in enumerate(rows):
        row = 5 + i
        pct_rev = round(r["revenue_ars"] / total_rev * 100, 1) if total_rev else 0
        pct_ord = round(r["ordenes"] / total_ord * 100, 1) if total_ord else 0
        vals = [r["categoria"], r["ordenes"], r["unidades_vendidas"],
                r["revenue_ars"], r["productos_distintos"],
                r["ticket_promedio_unidad"], pct_rev, pct_ord]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fills[i % 2]
            if col in (4, 6):
                cell.number_format = '#,##0.00'
            if col in (7, 8):
                cell.number_format = '0.0"%"'

    # Fila total
    tot_row = 5 + len(rows)
    tot_vals = ["TOTAL", total_ord, sum(r["unidades_vendidas"] for r in rows),
                total_rev, "", "", 100.0, 100.0]
    tot_fill = PatternFill("solid", fgColor="F2F2CC")
    for col, v in enumerate(tot_vals, 1):
        cell = ws.cell(row=tot_row, column=col, value=v)
        cell.font = Font(bold=True)
        cell.fill = tot_fill
        if col == 4:
            cell.number_format = '#,##0.00'

    # Anchos
    widths = [30, 10, 18, 18, 18, 24, 12, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    out = f"/Users/matiaslerer/Downloads/categorias_pret_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(out)
    return out


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Si pasaste el path del JSON como argumento, usalo
    if len(sys.argv) > 1:
        GOOGLE_SA_JSON = sys.argv[1]

    # Si no está seteado, preguntar
    if not GOOGLE_SA_JSON:
        GOOGLE_SA_JSON = input("Path al JSON de Google Service Account (Enter para crear Excel local): ").strip()

    orders   = fetch_orders()
    products = fetch_product_categories()
    rows     = analyze_by_category(orders, products)

    print("📊 Resultados por categoría:")
    print(f"{'Categoría':<35} {'Órdenes':>8} {'Unidades':>10} {'Revenue ARS':>15}")
    print("-" * 70)
    for r in rows:
        print(f"{r['categoria']:<35} {r['ordenes']:>8} {r['unidades_vendidas']:>10} {r['revenue_ars']:>15,.0f}")

    print()

    if GOOGLE_SA_JSON:
        print("📤 Subiendo a Google Sheets ...")
        try:
            url = build_google_sheet(rows, GOOGLE_SA_JSON, len(orders), DATE_FROM)
            print(f"\n✅ Google Sheet creado:")
            print(f"   {url}\n")
        except Exception as e:
            print(f"⚠️  Error con Google Sheets: {e}")
            print("   Generando Excel local como fallback ...")
            out = build_excel_fallback(rows, len(orders), DATE_FROM)
            print(f"✅ Excel guardado en: {out}")
    else:
        print("📁 Generando Excel local ...")
        out = build_excel_fallback(rows, len(orders), DATE_FROM)
        print(f"✅ Excel guardado en: {out}")
