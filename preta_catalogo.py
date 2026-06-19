#!/usr/bin/env python3
"""
PRET A HOME — Catálogo de Productos desde Tiendanube
=====================================================
Conecta a la API de Tiendanube, trae todos los productos con
variantes, stock por color/talle, precios y links, y genera
un Excel visualmente atractivo.

Requisitos:
    pip install requests openpyxl pillow

Uso:
    python preta_catalogo.py
"""

import requests
import time
import json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import sys

# ── CREDENCIALES ──────────────────────────────────────────────────
STORE_ID    = "2625285"
ACCESS_TOKEN = "7bf4cde46764d96772079d8cb1d10cd644aa35a0"
BASE_URL    = f"https://api.tiendanube.com/v1/{STORE_ID}"
HEADERS     = {
    "Authentication": f"bearer {ACCESS_TOKEN}",
    "User-Agent": "PretaHome-Catalogo/1.0 (reporte@pretahome.com)",
    "Content-Type": "application/json",
}
STORE_URL   = "https://www.pretahome.com/productos"

# ── PALETA PRET A HOME ────────────────────────────────────────────
C = {
    "nude":       "C4A882",
    "nude_dark":  "A0845C",
    "nude_soft":  "EDE0CE",
    "nude_xsoft": "F7F3ED",
    "black":      "1A1A1A",
    "white":      "FFFFFF",
    "off_white":  "F7F5F2",
    "grey_med":   "6B6B6B",
    "grey_lt":    "D9D4CE",
    "grey_xlt":   "F0EDE8",
    "green":      "4CAF50",
    "yellow":     "FFC107",
    "red":        "E53935",
    "green_bg":   "E8F5E9",
    "yellow_bg":  "FFF8E1",
    "red_bg":     "FFEBEE",
}

def hex_fill(hex_color, fill_type="solid"):
    return PatternFill(fill_type=fill_type, start_color=hex_color, end_color=hex_color)

def border(style="thin", color="D9D4CE"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(style="thin", color="D9D4CE"):
    return Border(bottom=Side(style=style, color=color))

# ── FETCH TIENDANUBE ──────────────────────────────────────────────
def fetch_all_products():
    products = []
    page = 1
    print("🔗 Conectando a Tiendanube API...")
    while True:
        url = f"{BASE_URL}/products?per_page=200&page={page}&fields=id,name,variants,categories,handle,published,updated_at"
        print(f"   → Página {page}...", end=" ", flush=True)
        resp = requests.get(url, headers=HEADERS, timeout=30)
        if resp.status_code == 401:
            print("\n❌ Token expirado o inválido. Verificá las credenciales.")
            sys.exit(1)
        if resp.status_code == 429:
            print(" (rate limit, esperando 5s...)", end=" ")
            time.sleep(5)
            continue
        if resp.status_code != 200:
            print(f"\n❌ Error {resp.status_code}: {resp.text[:200]}")
            sys.exit(1)
        batch = resp.json()
        if not batch:
            print("fin.")
            break
        products.extend(batch)
        print(f"{len(batch)} productos ({len(products)} total)")
        if len(batch) < 200:
            break
        page += 1
        time.sleep(0.5)
    print(f"\n✅ Total productos obtenidos: {len(products)}")
    return products

def fetch_categories():
    resp = requests.get(f"{BASE_URL}/categories?per_page=200", headers=HEADERS, timeout=30)
    if resp.status_code == 200:
        cats = {str(c["id"]): c.get("name", {}).get("es", "") for c in resp.json()}
        return cats
    return {}

# ── PROCESAR DATOS ────────────────────────────────────────────────
def parse_name(name_field):
    if isinstance(name_field, dict):
        return name_field.get("es") or name_field.get("pt") or list(name_field.values())[0] or ""
    return str(name_field or "")

def stock_status(stock, stock_management):
    if not stock_management:
        return "Sin control", C["grey_med"], C["grey_xlt"]
    if stock is None:
        return "Sin control", C["grey_med"], C["grey_xlt"]
    if stock <= 0:
        return "Sin stock", C["red"], C["red_bg"]
    if stock <= 5:
        return "Stock bajo", C["yellow"], C["yellow_bg"]
    return "OK", C["green"], C["green_bg"]

def build_rows(products, categories):
    rows = []
    for p in products:
        prod_name = parse_name(p.get("name", ""))
        handle    = p.get("handle", {})
        if isinstance(handle, dict):
            handle = handle.get("es") or list(handle.values())[0] or ""
        prod_url  = f"https://www.pretahome.com/productos/{handle}/"
        published = "Sí" if p.get("published") else "No"

        # Categorías
        cat_ids   = [str(c["id"]) for c in p.get("categories", [])]
        cat_names = " / ".join(filter(None, [categories.get(cid, "") for cid in cat_ids]))

        variants  = p.get("variants", [])
        if not variants:
            rows.append({
                "id": p["id"], "producto": prod_name, "categoria": cat_names,
                "sku": "", "color": "", "talle": "",
                "precio_lista": None, "precio_promo": None, "precio_revendedor": None,
                "stock": None, "stock_status": "Sin variantes",
                "stock_color": C["grey_med"], "stock_bg": C["grey_xlt"],
                "publicado": published, "url": prod_url,
            })
            continue

        for v in variants:
            # Atributos (color, talle, etc.)
            values = v.get("values", [])
            color_val = ""
            talle_val = ""
            for val in values:
                name_attr = parse_name(val.get("name", ""))
                name_lower = name_attr.lower()
                if any(x in name_lower for x in ["color", "colour", "color"]):
                    color_val = parse_name(val.get("es", val.get("name", "")))
                elif any(x in name_lower for x in ["talle", "size", "tamaño", "medida"]):
                    talle_val = parse_name(val.get("es", val.get("name", "")))
                else:
                    # Si solo hay un atributo, asignarlo a color
                    if not color_val:
                        color_val = parse_name(val.get("es", val.get("name", "")))

            # Precios
            precio_lista = None
            precio_promo = None
            try:
                if v.get("price") is not None:
                    precio_lista = float(v["price"])
                if v.get("promotional_price") is not None:
                    precio_promo = float(v["promotional_price"])
                    if precio_promo == precio_lista:
                        precio_promo = None
            except (ValueError, TypeError):
                pass

            precio_rev = round(precio_lista * 0.5, 2) if precio_lista else None

            # Stock
            stock_mgmt = v.get("stock_management", False)
            stock_qty  = v.get("stock", None)
            if isinstance(stock_qty, str):
                try: stock_qty = int(stock_qty)
                except: stock_qty = None

            status, s_color, s_bg = stock_status(stock_qty, stock_mgmt)
            sku = v.get("sku", "") or ""

            rows.append({
                "id":              p["id"],
                "producto":        prod_name,
                "categoria":       cat_names,
                "sku":             sku,
                "color":           color_val,
                "talle":           talle_val,
                "precio_lista":    precio_lista,
                "precio_promo":    precio_promo,
                "precio_revendedor": precio_rev,
                "stock":           stock_qty,
                "stock_status":    status,
                "stock_color":     s_color,
                "stock_bg":        s_bg,
                "publicado":       published,
                "url":             prod_url,
            })
    return rows

# ── BUILD EXCEL ───────────────────────────────────────────────────
def build_excel(rows, output_path):
    wb = Workbook()

    build_catalogo_sheet(wb, rows)
    build_resumen_sheet(wb, rows)

    # Eliminar hoja default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"\n💾 Excel guardado en: {output_path}")

# ── HOJA 1: CATÁLOGO COMPLETO ─────────────────────────────────────
def build_catalogo_sheet(wb, rows):
    ws = wb.create_sheet("📦 Catálogo Completo")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    # ── HEADER DECORATIVO ─────────────────────────────────────────
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 8

    # Merge y fondo oscuro en fila 2
    ws.merge_cells("A2:O2")
    ws["A2"] = "PRET A HOME  —  Catálogo de Productos"
    ws["A2"].font = Font(name="Calibri", size=20, bold=True, color=C["white"])
    ws["A2"].fill = hex_fill(C["black"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Franja nude fila 3
    ws.merge_cells("A3:O3")
    ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Tienda: pretahome.com  |  Precio revendedor = 50% OFF precio de lista"
    ws["A3"].font = Font(name="Calibri", size=9, color=C["black"])
    ws["A3"].fill = hex_fill(C["nude_soft"])
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    # Fila separadora
    for col in range(1, 16):
        ws.cell(row=4, column=col).fill = hex_fill(C["nude"])

    # ── CABECERAS ─────────────────────────────────────────────────
    headers = [
        ("PRODUCTO",          28),
        ("CATEGORÍA",         20),
        ("SKU",               14),
        ("COLOR / VARIANTE",  22),
        ("TALLE / MEDIDA",    18),
        ("PRECIO LISTA",      16),
        ("PRECIO PROMO",      16),
        ("COSTO REVENDEDOR",  18),
        ("STOCK",             10),
        ("ESTADO STOCK",      14),
        ("PUBLICADO",         11),
        ("URL TIENDANUBE",    50),
    ]

    COL_OFFSET = 2  # empezar en col B

    for col_i, (label, width) in enumerate(headers):
        col = col_i + COL_OFFSET
        cell = ws.cell(row=5, column=col)
        cell.value = label
        cell.font = Font(name="Calibri", size=9, bold=True, color=C["white"])
        cell.fill = hex_fill(C["black"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color=C["nude"]),
            right=Side(style="thin", color="3A3A3A"),
        )
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[5].height = 28

    # Col A: margen visual
    ws.column_dimensions["A"].width = 2

    # ── DATOS ─────────────────────────────────────────────────────
    # Agrupar por producto para filas alternadas
    prod_groups = {}
    for r in rows:
        pid = r["id"]
        prod_groups.setdefault(pid, []).append(r)

    ROW_START = 6
    cur_row = ROW_START
    alt = False

    for pid, group_rows in prod_groups.items():
        for i, r in enumerate(group_rows):
            bg = C["off_white"] if alt else C["white"]
            is_first = (i == 0)

            def cell_style(col, value, fmt=None, align="left", link=None, wrap=False):
                c = ws.cell(row=cur_row, column=col + COL_OFFSET - 1)
                c.value = value
                c.font = Font(name="Calibri", size=9, color=C["black"])
                c.fill = hex_fill(bg)
                c.alignment = Alignment(
                    horizontal=align, vertical="center",
                    wrap_text=wrap
                )
                c.border = Border(
                    bottom=Side(style="hair", color=C["grey_lt"]),
                    right=Side(style="hair", color=C["grey_lt"]),
                )
                if fmt:
                    c.number_format = fmt
                if link:
                    c.hyperlink = link
                    c.font = Font(name="Calibri", size=9, color="1155CC", underline="single")
                return c

            # Col 1: Producto (solo en primera variante del grupo)
            c1 = ws.cell(row=cur_row, column=COL_OFFSET)
            c1.value = r["producto"] if is_first else ""
            c1.font = Font(name="Calibri", size=9,
                           bold=is_first, color=C["black"])
            c1.fill = hex_fill(C["nude_xsoft"] if is_first else bg)
            c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c1.border = Border(
                left=Side(style="medium" if is_first else "hair", color=C["nude"] if is_first else C["grey_lt"]),
                bottom=Side(style="hair", color=C["grey_lt"]),
                right=Side(style="hair", color=C["grey_lt"]),
            )

            cell_style(2, r["categoria"], align="left")
            cell_style(3, r["sku"], align="center")
            cell_style(4, r["color"], align="left")
            cell_style(5, r["talle"], align="center")

            # Precios
            c_lista = cell_style(6, r["precio_lista"], fmt='$#,##0.00', align="right")
            c_promo = cell_style(7, r["precio_promo"], fmt='$#,##0.00', align="right")
            c_rev   = ws.cell(row=cur_row, column=COL_OFFSET + 7)
            c_rev.value = r["precio_revendedor"]
            c_rev.number_format = '$#,##0.00'
            c_rev.font = Font(name="Calibri", size=9, bold=True, color=C["nude_dark"])
            c_rev.fill = hex_fill(C["nude_soft"])
            c_rev.alignment = Alignment(horizontal="right", vertical="center")
            c_rev.border = Border(
                bottom=Side(style="hair", color=C["grey_lt"]),
                right=Side(style="hair", color=C["grey_lt"]),
            )

            # Stock
            c_stock = ws.cell(row=cur_row, column=COL_OFFSET + 8)
            c_stock.value = r["stock"]
            c_stock.font = Font(name="Calibri", size=9, bold=True, color=C["black"])
            c_stock.fill = hex_fill(r["stock_bg"])
            c_stock.alignment = Alignment(horizontal="center", vertical="center")
            c_stock.number_format = '#,##0'
            c_stock.border = Border(
                bottom=Side(style="hair", color=C["grey_lt"]),
                right=Side(style="hair", color=C["grey_lt"]),
            )

            # Estado stock (badge)
            c_est = ws.cell(row=cur_row, column=COL_OFFSET + 9)
            c_est.value = r["stock_status"]
            c_est.font = Font(name="Calibri", size=8, bold=True, color=r["stock_color"])
            c_est.fill = hex_fill(r["stock_bg"])
            c_est.alignment = Alignment(horizontal="center", vertical="center")
            c_est.border = Border(
                bottom=Side(style="hair", color=C["grey_lt"]),
                right=Side(style="hair", color=C["grey_lt"]),
            )

            # Publicado
            c_pub = ws.cell(row=cur_row, column=COL_OFFSET + 10)
            c_pub.value = r["publicado"]
            pub_color = C["green"] if r["publicado"] == "Sí" else C["red"]
            pub_bg    = C["green_bg"] if r["publicado"] == "Sí" else C["red_bg"]
            c_pub.font = Font(name="Calibri", size=8, bold=True, color=pub_color)
            c_pub.fill = hex_fill(pub_bg)
            c_pub.alignment = Alignment(horizontal="center", vertical="center")
            c_pub.border = Border(
                bottom=Side(style="hair", color=C["grey_lt"]),
                right=Side(style="hair", color=C["grey_lt"]),
            )

            # URL con hyperlink
            c_url = ws.cell(row=cur_row, column=COL_OFFSET + 11)
            c_url.value = r["url"]
            c_url.hyperlink = r["url"]
            c_url.font = Font(name="Calibri", size=8, color=C["nude_dark"], underline="single")
            c_url.fill = hex_fill(bg)
            c_url.alignment = Alignment(horizontal="left", vertical="center")
            c_url.border = Border(
                bottom=Side(style="hair", color=C["grey_lt"]),
                right=Side(style="hair", color=C["grey_lt"]),
            )

            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        # Línea divisora entre productos
        for col in range(COL_OFFSET, COL_OFFSET + 12):
            ws.cell(row=cur_row - 1, column=col).border = Border(
                bottom=Side(style="thin", color=C["grey_lt"]),
                right=Side(style="hair", color=C["grey_lt"]),
            )

        alt = not alt

    # ── PIE DE PÁGINA ─────────────────────────────────────────────
    cur_row += 1
    ws.merge_cells(f"B{cur_row}:M{cur_row}")
    ws.cell(row=cur_row, column=2).value = (
        f"Total filas: {len(rows)}  |  "
        f"Fuente: Tiendanube API — api.tiendanube.com  |  "
        f"Pret a Home Store ID: {STORE_ID}"
    )
    ws.cell(row=cur_row, column=2).font = Font(name="Calibri", size=8, italic=True, color=C["grey_med"])
    ws.cell(row=cur_row, column=2).fill = hex_fill(C["off_white"])

    # ── AUTOFILTER ────────────────────────────────────────────────
    last_col = get_column_letter(COL_OFFSET + 11)
    ws.auto_filter.ref = f"B5:{last_col}{cur_row - 2}"

    # ── INMOVILIZAR ───────────────────────────────────────────────
    ws.freeze_panes = f"B6"

    # ── ZOOM ──────────────────────────────────────────────────────
    ws.sheet_view.zoomScale = 90

# ── HOJA 2: RESUMEN POR CATEGORÍA ────────────────────────────────
def build_resumen_sheet(wb, rows):
    ws = wb.create_sheet("📊 Resumen por Categoría", 0)  # primera hoja
    ws.sheet_view.showGridLines = False

    # Header
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 44
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 8

    ws.merge_cells("A2:K2")
    ws["A2"] = "PRET A HOME  —  Resumen de Catálogo por Categoría"
    ws["A2"].font = Font(name="Calibri", size=22, bold=True, color=C["white"])
    ws["A2"].fill = hex_fill(C["black"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:K3")
    ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  pretahome.com  |  Datos en tiempo real desde Tiendanube"
    ws["A3"].font = Font(name="Calibri", size=9, color=C["black"])
    ws["A3"].fill = hex_fill(C["nude_soft"])
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 12):
        ws.cell(row=4, column=col).fill = hex_fill(C["nude"])

    # Ancho columnas
    col_widths = [2, 28, 14, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    # ── KPI CARDS ──────────────────────────────────────────────────
    # Calcular KPIs globales
    total_prods   = len(set(r["id"] for r in rows))
    total_vars    = len(rows)
    sin_stock     = sum(1 for r in rows if r["stock"] == 0)
    stock_bajo    = sum(1 for r in rows if r["stock"] is not None and 0 < r["stock"] <= 5)
    publicados    = sum(1 for r in rows if r["publicado"] == "Sí")

    kpis = [
        ("PRODUCTOS",        total_prods,  C["black"],     C["nude_soft"]),
        ("VARIANTES",        total_vars,   C["black"],     C["off_white"]),
        ("PUBLICADOS",       publicados,   C["green"],     C["green_bg"]),
        ("SIN STOCK",        sin_stock,    C["red"],       C["red_bg"]),
        ("STOCK BAJO",       stock_bajo,   "B8860B",       C["yellow_bg"]),
    ]

    kpi_row_label = 6
    kpi_row_value = 7
    kpi_start_col = 2
    ws.row_dimensions[5].height = 10
    ws.row_dimensions[kpi_row_label].height = 18
    ws.row_dimensions[kpi_row_value].height = 40
    ws.row_dimensions[8].height = 10

    for i, (label, value, fg, bg) in enumerate(kpis):
        col = kpi_start_col + i * 2
        # Merge 2 celdas para cada KPI
        ws.merge_cells(start_row=kpi_row_label, start_column=col, end_row=kpi_row_label, end_column=col+1)
        ws.merge_cells(start_row=kpi_row_value, start_column=col, end_row=kpi_row_value, end_column=col+1)

        c_label = ws.cell(row=kpi_row_label, column=col)
        c_label.value = label
        c_label.font = Font(name="Calibri", size=8, bold=True, color=C["grey_med"])
        c_label.fill = hex_fill(bg)
        c_label.alignment = Alignment(horizontal="center", vertical="bottom")

        c_val = ws.cell(row=kpi_row_value, column=col)
        c_val.value = value
        c_val.font = Font(name="Calibri", size=28, bold=True, color=fg)
        c_val.fill = hex_fill(bg)
        c_val.alignment = Alignment(horizontal="center", vertical="center")

        # Borde
        for r in [kpi_row_label, kpi_row_value]:
            for c in [col, col+1]:
                cell = ws.cell(row=r, column=c)
                cell.border = Border(
                    left=Side(style="medium" if c == col else "hair", color=C["grey_lt"]),
                    right=Side(style="medium" if c == col+1 else "hair", color=C["grey_lt"]),
                    top=Side(style="medium", color=C["grey_lt"]),
                    bottom=Side(style="medium", color=C["grey_lt"]),
                )

    # ── TABLA POR CATEGORÍA ────────────────────────────────────────
    # Calcular métricas por categoría
    from collections import defaultdict
    cat_data = defaultdict(lambda: {
        "productos": set(), "variantes": 0,
        "sin_stock": 0, "stock_bajo": 0, "ok": 0,
        "precio_min": None, "precio_max": None,
    })

    for r in rows:
        cat = r["categoria"] or "Sin categoría"
        cat_data[cat]["productos"].add(r["id"])
        cat_data[cat]["variantes"] += 1
        if r["stock"] == 0:
            cat_data[cat]["sin_stock"] += 1
        elif r["stock"] is not None and 0 < r["stock"] <= 5:
            cat_data[cat]["stock_bajo"] += 1
        else:
            cat_data[cat]["ok"] += 1
        if r["precio_lista"]:
            mn = cat_data[cat]["precio_min"]
            mx = cat_data[cat]["precio_max"]
            cat_data[cat]["precio_min"] = min(mn, r["precio_lista"]) if mn else r["precio_lista"]
            cat_data[cat]["precio_max"] = max(mx, r["precio_lista"]) if mx else r["precio_lista"]

    TABLE_START = 10
    ws.row_dimensions[9].height = 10

    # Cabeceras tabla categoría
    cat_headers = [
        ("CATEGORÍA",        26),
        ("PRODUCTOS",        12),
        ("VARIANTES",        12),
        ("PRECIO MÍNIMO",    16),
        ("PRECIO MÁXIMO",    16),
        ("COSTO REV. MÍN",   16),
        ("COSTO REV. MÁX",   16),
        ("SIN STOCK",        12),
        ("STOCK BAJO",       12),
        ("OK",               10),
    ]
    ws.row_dimensions[TABLE_START].height = 28

    for i, (label, width) in enumerate(cat_headers):
        col = i + 2
        cell = ws.cell(row=TABLE_START, column=col)
        cell.value = label
        cell.font = Font(name="Calibri", size=9, bold=True, color=C["white"])
        cell.fill = hex_fill(C["black"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color=C["nude"]),
            right=Side(style="thin", color="3A3A3A"),
        )
        ws.column_dimensions[get_column_letter(col)].width = width

    # Filas de categorías (ordenadas por variantes desc)
    sorted_cats = sorted(cat_data.items(), key=lambda x: -len(x[1]["productos"]))

    cur_row = TABLE_START + 1
    for alt_i, (cat_name, cd) in enumerate(sorted_cats):
        bg = C["off_white"] if alt_i % 2 == 0 else C["white"]
        ws.row_dimensions[cur_row].height = 18

        def cat_cell(col, value, fmt=None, fg=C["black"], cell_bg=None, bold=False):
            c = ws.cell(row=cur_row, column=col)
            c.value = value
            c.font = Font(name="Calibri", size=9, bold=bold, color=fg)
            c.fill = hex_fill(cell_bg or bg)
            c.alignment = Alignment(horizontal="center" if col > 2 else "left",
                                    vertical="center")
            c.border = Border(
                bottom=Side(style="hair", color=C["grey_lt"]),
                right=Side(style="hair", color=C["grey_lt"]),
            )
            if fmt:
                c.number_format = fmt
            return c

        cat_cell(2, cat_name, bold=True)
        cat_cell(3, len(cd["productos"]))
        cat_cell(4, cd["variantes"])
        cat_cell(5, cd["precio_min"], fmt='$#,##0')
        cat_cell(6, cd["precio_max"], fmt='$#,##0')
        cat_cell(7, cd["precio_min"]*0.5 if cd["precio_min"] else None, fmt='$#,##0',
                 fg=C["nude_dark"], cell_bg=C["nude_soft"], bold=True)
        cat_cell(8, cd["precio_max"]*0.5 if cd["precio_max"] else None, fmt='$#,##0',
                 fg=C["nude_dark"], cell_bg=C["nude_soft"], bold=True)

        # Sin stock
        ss = cd["sin_stock"]
        cat_cell(9, ss if ss > 0 else "-",
                 fg=C["red"] if ss > 0 else C["grey_med"],
                 cell_bg=C["red_bg"] if ss > 0 else bg,
                 bold=(ss > 0))

        # Stock bajo
        sb = cd["stock_bajo"]
        cat_cell(10, sb if sb > 0 else "-",
                 fg="B8860B" if sb > 0 else C["grey_med"],
                 cell_bg=C["yellow_bg"] if sb > 0 else bg,
                 bold=(sb > 0))

        # OK
        ok = cd["ok"]
        cat_cell(11, ok if ok > 0 else "-",
                 fg=C["green"] if ok > 0 else C["grey_med"],
                 cell_bg=C["green_bg"] if ok > 0 else bg,
                 bold=(ok > 0))

        cur_row += 1

    # Fila de totales
    ws.row_dimensions[cur_row].height = 22
    total_row_data = [
        ("TOTAL", None, None),
        (f"=SUM(C{TABLE_START+1}:C{cur_row-1})", None, None),
        (f"=SUM(D{TABLE_START+1}:D{cur_row-1})", None, None),
        (None, '$#,##0', None),
        (None, '$#,##0', None),
        (None, '$#,##0', None),
        (None, '$#,##0', None),
        (f"=SUM(I{TABLE_START+1}:I{cur_row-1})", None, C["red_bg"]),
        (f"=SUM(J{TABLE_START+1}:J{cur_row-1})", None, C["yellow_bg"]),
        (f"=SUM(K{TABLE_START+1}:K{cur_row-1})", None, C["green_bg"]),
    ]
    for i, (val, fmt, cell_bg) in enumerate(total_row_data):
        col = i + 2
        c = ws.cell(row=cur_row, column=col)
        c.value = val
        c.font = Font(name="Calibri", size=9, bold=True, color=C["white"])
        c.fill = hex_fill(C["black"])
        c.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")
        c.border = Border(
            top=Side(style="medium", color=C["nude"]),
            bottom=Side(style="medium", color=C["nude"]),
            right=Side(style="hair", color="3A3A3A"),
        )
        if fmt:
            c.number_format = fmt

    # ── PIE ────────────────────────────────────────────────────────
    cur_row += 2
    ws.merge_cells(f"B{cur_row}:K{cur_row}")
    ws.cell(row=cur_row, column=2).value = (
        "ℹ️  Los precios son en pesos argentinos (ARS).  "
        "El precio revendedor es 50% OFF del precio de lista web.  "
        "El stock se actualiza en tiempo real desde Tiendanube."
    )
    ws.cell(row=cur_row, column=2).font = Font(name="Calibri", size=8, italic=True, color=C["grey_med"])
    ws.cell(row=cur_row, column=2).fill = hex_fill(C["off_white"])
    ws.cell(row=cur_row, column=2).alignment = Alignment(horizontal="center")

    ws.auto_filter.ref = f"B{TABLE_START}:K{cur_row-3}"
    ws.freeze_panes = f"B{TABLE_START+1}"
    ws.sheet_view.zoomScale = 95

# ── MAIN ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  PRET A HOME — Catálogo Tiendanube → Excel")
    print("=" * 55)

    products   = fetch_all_products()
    categories = fetch_categories()

    print(f"\n📂 Categorías encontradas: {len(categories)}")
    print("🔧 Procesando variantes y precios...")

    rows = build_rows(products, categories)
    print(f"📋 Total filas (variantes): {len(rows)}")

    output = "PretaHome_Catalogo.xlsx"
    print(f"\n📊 Generando Excel visualmente atractivo...")
    build_excel(rows, output)

    print("\n✅ ¡Listo! Abrí el archivo:")
    print(f"   📁 {output}")
    print("\n   Hoja 1 — 📊 Resumen por Categoría")
    print("   Hoja 2 — 📦 Catálogo Completo (con filtros y links)")
    print()
