#!/usr/bin/env python3
"""
generar_reporte.py — Reporte de ventas
Pret a Home + Casa Lavan | TN + ML → Excel

Uso:
    python3 generar_reporte.py
    python3 generar_reporte.py --days 30
    python3 generar_reporte.py --config /ruta/a/config.json
    python3 generar_reporte.py --coverage 14

Requiere:
    pip install requests openpyxl
"""

import json, requests, time, threading, argparse, sys, re
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from pathlib import Path

# ─── ARGS ─────────────────────────────────────────────────────────────────────

parser = argparse.ArgumentParser()
parser.add_argument("--days",     type=int, default=60)
parser.add_argument("--config",   type=str, default=None)
parser.add_argument("--output",   type=str, default=None)
parser.add_argument("--coverage", type=int, default=10,
                    help="Días de cobertura target para cálculo de reposición (default: 10)")
parser.add_argument("--suppliers-url", type=str, default=None,
                    help="URL del Google Sheet de SKU→Proveedor (override del default hardcodeado)")
args = parser.parse_args()

DAYS          = args.days
COVERAGE_DAYS = args.coverage
NOW           = datetime.now(timezone.utc)
DATE_FROM     = NOW - timedelta(days=DAYS)
DATE_FROM_ISO = DATE_FROM.strftime("%Y-%m-%dT%H:%M:%SZ")
DATE_FROM_STR = DATE_FROM.strftime("%Y-%m-%d")
DATE_TO_STR   = NOW.strftime("%Y-%m-%d")

# ─── PROVEEDORES ──────────────────────────────────────────────────────────────
# Prioridad: 1) mapeo explícito SKU→Proveedor desde Google Sheets
#            2) detección por keywords en nombre del producto (fallback)
#
# Para actualizar el mapeo, editá el Google Sheet — no hace falta tocar el script.
# Para usar otro Sheet pasá: --suppliers-url "https://docs.google.com/spreadsheets/d/SHEET_ID/..."

import csv, io

_DEFAULT_SUPPLIERS_SHEET = (
    "https://docs.google.com/spreadsheets/d/"
    "1_BPBfZ-GdoHwL-vS2rdIfFpYGRdbhzsqNlFmJ1VwLMI"
    "/export?format=csv&gid=0"
)

SUPPLIER_KEYWORDS = [
    ("Home Concept",   ["home concept"]),
    ("Jean Cartier",   ["jean cartier"]),
    ("Franco Valente", ["franco valente"]),
    ("City Blanco",    ["city blanco"]),
    ("Alcoyana",       ["alcoyana"]),
    ("Palette",        ["palette"]),
    ("Sakura",         ["sakura"]),
    ("Pret",           ["pret a home", "pret"]),
]
SUPPLIER_OTHER = "Otros"

def _load_sku_supplier_map(url: str) -> dict:
    """Descarga el Google Sheet como CSV y devuelve {sku_upper: proveedor}."""
    mapping = {}
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        reader = csv.DictReader(io.StringIO(r.text))
        for row in reader:
            sku = (row.get("SKU") or "").strip()
            proveedor = (row.get("Proveedor") or "").strip()
            if sku and proveedor:
                mapping[sku.upper()] = proveedor.strip().title()
        print(f"  ✓ Mapeo SKU→Proveedor cargado: {len(mapping)} SKUs desde Google Sheets")
    except Exception as e:
        print(f"  ⚠ No se pudo cargar el sheet de proveedores ({e}). Se usará solo detección por keywords.")
    return mapping

_suppliers_url = getattr(args, "suppliers_url", None) or _DEFAULT_SUPPLIERS_SHEET
SKU_SUPPLIER_MAP = _load_sku_supplier_map(_suppliers_url)

def detect_supplier(product_name: str, sku: str = "") -> str:
    """Primero busca el SKU en el mapeo explícito; si no está, cae a keywords."""
    if sku:
        result = SKU_SUPPLIER_MAP.get(sku.upper())
        if result:
            return result
    name_lc = product_name.lower()
    for supplier, keywords in SUPPLIER_KEYWORDS:
        if any(kw in name_lc for kw in keywords):
            return supplier
    return SUPPLIER_OTHER

# ─── CONFIG ───────────────────────────────────────────────────────────────────

def find_config():
    if args.config:
        p = Path(args.config)
        if p.exists():
            return p
        sys.exit(f"Config no encontrado: {args.config}")
    candidates = [
        Path(__file__).parent / "config.json",
        Path.home() / "Desktop" / "ecommerce-bi" / "config.json",
        Path.home() / "Downloads" / "ecommerce-bi" / "config.json",
    ]
    for p in candidates:
        if p.exists():
            return p
    sys.exit("No se encontró config.json. Pasalo con --config /ruta/config.json")

config_path = find_config()
with open(config_path) as f:
    CONFIG = json.load(f)

channels = CONFIG["channels"]
output_path = args.output or str(config_path.parent / f"reporte_ventas_{DATE_TO_STR}.xlsx")

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def tnlog(msg): print(f"  {msg}", flush=True)

# ─── ML AUTH ──────────────────────────────────────────────────────────────────

_ml_tokens = {}

def ml_ensure_token(key, cfg):
    if key in _ml_tokens:
        return _ml_tokens[key]
    token = cfg.get("access_token", "")
    refresh = cfg.get("refresh_token")
    if refresh:
        r = requests.post("https://api.mercadolibre.com/oauth/token", data={
            "grant_type":    "refresh_token",
            "client_id":     cfg["client_id"],
            "client_secret": cfg["client_secret"],
            "refresh_token": refresh,
        }, timeout=15)
        if r.status_code == 200:
            token = r.json()["access_token"]
            tnlog(f"✓ Token ML renovado para {cfg['label']}")
        else:
            tnlog(f"⚠ No se pudo renovar token ML {cfg['label']}, usando el guardado")
    _ml_tokens[key] = token
    return token

def ml_get(url, token, params=None, retries=5):
    headers = {"Authorization": f"Bearer {token}"}
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=20)
            if r.status_code == 200:
                return r.json()
            if r.status_code == 429:
                time.sleep(5 * (attempt + 1))
                continue
            if r.status_code == 500:
                time.sleep(3 * (attempt + 1))
                continue
            if r.status_code == 401:
                return None
        except requests.exceptions.Timeout:
            time.sleep(3 * (attempt + 1))
            continue
        except Exception:
            time.sleep(2)
            continue
    return None

# ─── TN FETCH ─────────────────────────────────────────────────────────────────

def tn_headers(cfg):
    return {
        "Authentication": f"bearer {cfg['access_token']}",
        "User-Agent":     "EcommerceBi (bi@pretahome.com)",
    }

def tn_get_orders(cfg):
    base = f"https://api.tiendanube.com/v1/{cfg['store_id']}"
    orders, page = [], 1
    while True:
        r = requests.get(f"{base}/orders", headers=tn_headers(cfg), params={
            "per_page": 200, "page": page,
            "created_at_min": DATE_FROM_ISO,
            "fields": "id,created_at,status,payment_status,products",
        }, timeout=20)
        if r.status_code != 200:
            tnlog(f"⚠ TN orders error {r.status_code} [{cfg['label']}]: {r.text[:300]}")
            break
        batch = r.json()
        if not batch:
            break
        for o in batch:
            if o.get("payment_status") in ("paid", "authorized") and o.get("status") != "cancelled":
                orders.append(o)
        if len(batch) < 200:
            break
        page += 1
        time.sleep(0.5)
    return orders

def tn_get_products(cfg):
    base = f"https://api.tiendanube.com/v1/{cfg['store_id']}"
    products, page = [], 1
    while True:
        r = requests.get(f"{base}/products", headers=tn_headers(cfg), params={
            "per_page": 200, "page": page,
        }, timeout=30)
        if r.status_code != 200:
            tnlog(f"⚠ TN products error {r.status_code} [{cfg['label']}]: {r.text[:200]}")
            break
        batch = r.json()
        if not batch:
            break
        products.extend(batch)
        tnlog(f"  TN products page {page}: {len(batch)} productos [{cfg['label']}]")
        if len(batch) < 200:
            break
        page += 1
        time.sleep(0.5)
    return products

def parse_tn_sales(orders):
    """Returns (sales_total, sales_first_half, sales_second_half) dicts keyed by sku."""
    mid = DATE_FROM + (NOW - DATE_FROM) / 2
    sales_total  = defaultdict(int)
    sales_first  = defaultdict(int)
    sales_second = defaultdict(int)
    for o in orders:
        try:
            order_date = datetime.fromisoformat(
                o.get("created_at", "").replace("Z", "+00:00")
            )
        except Exception:
            order_date = NOW
        for item in o.get("products", []):
            sku = str(item.get("sku") or item.get("variant_id") or "")
            qty = item.get("quantity", 1)
            if sku:
                sales_total[sku]  += qty
                if order_date < mid:
                    sales_first[sku]  += qty
                else:
                    sales_second[sku] += qty
    return sales_total, sales_first, sales_second

# ─── ML FETCH ─────────────────────────────────────────────────────────────────

def ml_get_orders(key, cfg):
    token   = ml_ensure_token(key, cfg)
    user_id = cfg["user_id"]
    label   = cfg["label"]

    CHUNK_DAYS = 7
    ML_MAX     = 10000
    all_orders = []
    chunk_start = DATE_FROM

    while chunk_start < NOW:
        chunk_end = min(chunk_start + timedelta(days=CHUNK_DAYS), NOW)
        date_from_str = chunk_start.strftime("%Y-%m-%dT%H:%M:%S.000-00:00")
        date_to_str   = chunk_end.strftime("%Y-%m-%dT%H:%M:%S.000-00:00")

        probe = ml_get("https://api.mercadolibre.com/orders/search", token, params={
            "seller": user_id, "order.status": "paid",
            "order.date_created.from": date_from_str,
            "order.date_created.to":   date_to_str,
            "sort": "date_asc", "offset": 0, "limit": 1,
        })
        if not probe:
            chunk_start = chunk_end
            continue

        chunk_total = probe.get("paging", {}).get("total", 0)

        if chunk_total > ML_MAX:
            tnlog(f"  ⚠ {label}: chunk {chunk_start.strftime('%d/%m')}→{chunk_end.strftime('%d/%m')} tiene {chunk_total} órdenes, subdividiendo...")
            CHUNK_DAYS = max(1, CHUNK_DAYS // 2)
            continue

        MAX_CHUNK_RETRIES = 3
        for chunk_attempt in range(MAX_CHUNK_RETRIES):
            offset, limit = 0, 50
            chunk_orders = []
            chunk_ok = True
            while offset < chunk_total:
                data = ml_get("https://api.mercadolibre.com/orders/search", token, params={
                    "seller": user_id, "order.status": "paid",
                    "order.date_created.from": date_from_str,
                    "order.date_created.to":   date_to_str,
                    "sort": "date_asc", "offset": offset, "limit": limit,
                })
                if not data:
                    chunk_ok = False
                    break
                batch = data.get("results", [])
                if not batch:
                    break
                chunk_orders.extend(batch)
                if len(batch) < limit:
                    break
                offset += limit
                time.sleep(0.35)

            recovered = len(chunk_orders)
            if chunk_ok and recovered >= chunk_total * 0.98:
                break
            else:
                if chunk_attempt < MAX_CHUNK_RETRIES - 1:
                    tnlog(f"  ↩ {label}: chunk incompleto ({recovered}/{chunk_total}), reintentando...")
                    time.sleep(5)
                else:
                    tnlog(f"  ⚠ {label}: chunk incompleto tras {MAX_CHUNK_RETRIES} intentos ({recovered}/{chunk_total})")

        all_orders.extend(chunk_orders)
        tnlog(f"  ✓ {label}: {chunk_start.strftime('%d/%m')}→{chunk_end.strftime('%d/%m')} — {len(chunk_orders)}/{chunk_total} órdenes")
        chunk_start = chunk_end
        time.sleep(0.3)

    tnlog(f"  ✓ {label}: TOTAL {len(all_orders)} órdenes recuperadas en {DAYS} días")
    return all_orders, token

def parse_ml_sales(orders):
    """Returns (sales_total, sales_first_half, sales_second_half) dicts.

    Key strategy (fixes double-count bug):
    - Si el item tiene seller_sku → guardar SOLO como ("sku", sku)
    - Si NO tiene seller_sku    → guardar como ("id", item_id) para resolver después
    Nunca guardar ambos para el mismo item.
    """
    mid = DATE_FROM + (NOW - DATE_FROM) / 2
    sales_total  = defaultdict(int)
    sales_first  = defaultdict(int)
    sales_second = defaultdict(int)

    for o in orders:
        try:
            order_date = datetime.fromisoformat(
                o.get("date_created", "").replace("Z", "+00:00")
            )
        except Exception:
            order_date = NOW

        for item in o.get("order_items", []):
            item_obj = item.get("item", {})
            item_id  = str(item_obj.get("id", ""))
            # ML puede devolver el SKU en seller_sku o seller_custom_field
            sku = str(item_obj.get("seller_sku") or item_obj.get("seller_custom_field") or "").strip()
            qty = item.get("quantity", 1)

            # Usar SKU si existe, si no usar id (se resolverá a SKU más tarde via catálogo)
            if sku:
                key = ("sku", sku)
            elif item_id:
                key = ("id", item_id)
            else:
                continue  # sin identificador, ignorar

            sales_total[key] += qty
            if order_date < mid:
                sales_first[key]  += qty
            else:
                sales_second[key] += qty

    return sales_total, sales_first, sales_second

def ml_get_all_items(key, cfg):
    token   = ml_ensure_token(key, cfg)
    user_id = cfg["user_id"]
    item_ids, offset = [], 0
    while True:
        data = ml_get(f"https://api.mercadolibre.com/users/{user_id}/items/search",
                      token, params={"offset": offset, "limit": 100})
        if not data:
            break
        batch = data.get("results", [])
        if not batch:
            break
        item_ids.extend(batch)
        if len(batch) < 100:
            break
        offset += 100
        time.sleep(0.3)
    details = {}
    for i in range(0, len(item_ids), 20):
        batch_ids = item_ids[i:i+20]
        data = ml_get("https://api.mercadolibre.com/items", token,
                      params={"ids": ",".join(batch_ids)})
        if data:
            for entry in data:
                if entry.get("code") == 200:
                    body = entry["body"]
                    details[body["id"]] = body
        time.sleep(0.3)
    return details

# ─── THREAD WORKERS ───────────────────────────────────────────────────────────

results = {}
fetch_errors = {}

def worker_tn(key, cfg):
    label = cfg["label"]
    try:
        tnlog(f"→ {label}: obteniendo órdenes...")
        orders = tn_get_orders(cfg)
        tnlog(f"✓ {label}: {len(orders)} órdenes obtenidas")
        tnlog(f"→ {label}: obteniendo productos...")
        products = tn_get_products(cfg)
        tnlog(f"✓ {label}: {len(products)} productos obtenidos")
        sales_total, sales_first, sales_second = parse_tn_sales(orders)
        tnlog(f"✓ {label}: {len(sales_total)} SKUs con ventas")
        results[key] = {
            "orders": orders,
            "products": products,
            "sales": sales_total,
            "sales_first": sales_first,
            "sales_second": sales_second,
        }
        tnlog(f"✓ {label}: COMPLETO — {len(orders)} órdenes | {len(products)} productos | {len(sales_total)} SKUs con ventas")
    except Exception as e:
        import traceback
        fetch_errors[key] = str(e)
        results[key] = {"orders": [], "products": [], "sales": {},
                        "sales_first": {}, "sales_second": {}}
        tnlog(f"✗ {label}: ERROR — {e}")
        tnlog(f"  Traceback: {traceback.format_exc()}")

def worker_ml(key, cfg):
    label = cfg["label"]
    try:
        tnlog(f"→ {label}: obteniendo órdenes...")
        orders, token = ml_get_orders(key, cfg)
        sales_total, sales_first, sales_second = parse_ml_sales(orders)
        tnlog(f"→ {label}: obteniendo catálogo ML...")
        item_details = ml_get_all_items(key, cfg)
        results[key] = {
            "orders": orders,
            "sales": sales_total,
            "sales_first": sales_first,
            "sales_second": sales_second,
            "item_details": item_details,
        }
        tnlog(f"✓ {label}: {len(orders)} órdenes | {len(item_details)} items en catálogo")
    except Exception as e:
        fetch_errors[key] = str(e)
        results[key] = {"orders": [], "sales": defaultdict(int),
                        "sales_first": defaultdict(int), "sales_second": defaultdict(int),
                        "item_details": {}}
        tnlog(f"✗ {label}: {e}")

# ─── BUILD PRODUCT TABLE ──────────────────────────────────────────────────────

def get_name(product):
    n = product.get("name", {})
    if isinstance(n, dict):
        return n.get("es") or n.get("en") or next(iter(n.values()), "")
    return str(n)

def days_active(created_at_str):
    try:
        created = datetime.fromisoformat(created_at_str.replace("Z", "+00:00"))
        if created > NOW:
            return 0
        active_from = max(created, DATE_FROM)
        return max(1, (NOW - active_from).days)
    except Exception:
        return DAYS

def calc_trend(sold_first, sold_second, active_days):
    """Returns (trend_label, trend_pct) comparing 2nd half vs 1st half velocity."""
    half = max(active_days / 2, 1)
    vel1 = sold_first  / half
    vel2 = sold_second / half
    if vel1 == 0 and vel2 == 0:
        return "—", None
    if vel1 == 0:
        return "🚀 Nueva", None
    pct = round((vel2 - vel1) / vel1 * 100, 1)
    if pct >= 20:
        label = f"↑ +{pct:.0f}%"
    elif pct <= -20:
        label = f"↓ {pct:.0f}%"
    else:
        label = f"→ {pct:+.0f}%"
    return label, pct

def units_to_order(stock, vel_diaria, coverage=COVERAGE_DAYS):
    """Units needed to reach `coverage` days of stock at current velocity."""
    target = vel_diaria * coverage
    need   = max(0, round(target - stock))
    return need

def build_rows(results):
    tn_pret_products  = results.get("tn_pret", {}).get("products", [])
    tn_lavan_products = results.get("tn_lavan", {}).get("products", [])

    tn_pret_sales    = results.get("tn_pret", {}).get("sales", {})
    tn_lavan_sales   = results.get("tn_lavan", {}).get("sales", {})
    tn_pret_first    = results.get("tn_pret", {}).get("sales_first", {})
    tn_pret_second   = results.get("tn_pret", {}).get("sales_second", {})
    tn_lavan_first   = results.get("tn_lavan", {}).get("sales_first", {})
    tn_lavan_second  = results.get("tn_lavan", {}).get("sales_second", {})

    ml_pret_sales    = results.get("ml_pret", {}).get("sales", defaultdict(int))
    ml_lavan_sales   = results.get("ml_lavan", {}).get("sales", defaultdict(int))
    ml_pret_first    = results.get("ml_pret", {}).get("sales_first", defaultdict(int))
    ml_pret_second   = results.get("ml_pret", {}).get("sales_second", defaultdict(int))
    ml_lavan_first   = results.get("ml_lavan", {}).get("sales_first", defaultdict(int))
    ml_lavan_second  = results.get("ml_lavan", {}).get("sales_second", defaultdict(int))

    ml_pret_items    = results.get("ml_pret", {}).get("item_details", {})
    ml_lavan_items   = results.get("ml_lavan", {}).get("item_details", {})

    def ml_sales_by_sku(ml_sales):
        by_sku = defaultdict(int)
        for (kind, key), qty in ml_sales.items():
            if kind == "sku":
                by_sku[key] += qty
        return by_sku

    ml_pret_by_sku   = ml_sales_by_sku(ml_pret_sales)
    ml_lavan_by_sku  = ml_sales_by_sku(ml_lavan_sales)
    ml_pret_1h_sku   = ml_sales_by_sku(ml_pret_first)
    ml_pret_2h_sku   = ml_sales_by_sku(ml_pret_second)
    ml_lavan_1h_sku  = ml_sales_by_sku(ml_lavan_first)
    ml_lavan_2h_sku  = ml_sales_by_sku(ml_lavan_second)

    def build_id_to_sku(item_details):
        d = {}
        for item_id, body in item_details.items():
            sku = body.get("seller_custom_field") or body.get("seller_sku") or ""
            if sku:
                d[item_id] = sku
        return d

    def build_sku_to_permalink(item_details):
        d = {}
        for item_id, body in item_details.items():
            sku       = body.get("seller_custom_field") or body.get("seller_sku") or ""
            permalink = body.get("permalink") or body.get("secure_permalink") or ""
            if sku and permalink:
                d[sku] = permalink
            elif permalink:
                d[item_id] = permalink
        return d

    ml_pret_sku_to_url  = build_sku_to_permalink(ml_pret_items)
    ml_lavan_sku_to_url = build_sku_to_permalink(ml_lavan_items)

    def build_tn_url_map(products):
        d = {}
        for prod in products:
            pid = str(prod.get("id", ""))
            url = prod.get("canonical_url") or prod.get("seo_url") or ""
            if not url:
                handle = prod.get("handle", {})
                if isinstance(handle, dict):
                    handle = handle.get("es") or handle.get("en") or next(iter(handle.values()), "")
                if handle:
                    url = f"/{handle}"
            d[pid] = url
        return d

    tn_pret_url_map  = build_tn_url_map(tn_pret_products)
    tn_lavan_url_map = build_tn_url_map(tn_lavan_products)

    ml_pret_id_to_sku  = build_id_to_sku(ml_pret_items)
    ml_lavan_id_to_sku = build_id_to_sku(ml_lavan_items)

    # Resolver ventas guardadas como id → mapear a SKU (también halves)
    def resolve_id_sales(ml_sales, id_to_sku, by_sku, h1_sku, h2_sku, ml_first, ml_second):
        for (kind, k), qty in ml_sales.items():
            if kind == "id" and k in id_to_sku:
                mapped = id_to_sku[k]
                by_sku[mapped]  += qty
                h1_sku[mapped]  += ml_first.get(("id", k), 0)
                h2_sku[mapped]  += ml_second.get(("id", k), 0)

    resolve_id_sales(ml_pret_sales,  ml_pret_id_to_sku,  ml_pret_by_sku,
                     ml_pret_1h_sku,  ml_pret_2h_sku,  ml_pret_first,  ml_pret_second)
    resolve_id_sales(ml_lavan_sales, ml_lavan_id_to_sku, ml_lavan_by_sku,
                     ml_lavan_1h_sku, ml_lavan_2h_sku, ml_lavan_first, ml_lavan_second)

    # ── Construir mapa de SKU → ventas de TN Lavan (para sumar al row de TN Pret)
    # Índice: sku → (sold_tn_lavan, tn_lavan_1h, tn_lavan_2h, url_tn_lavan)
    tn_lavan_sku_index = {}
    for prod in tn_lavan_products:
        lavan_url_map = tn_lavan_url_map
        prod_id = str(prod.get("id", ""))
        lavan_url = lavan_url_map.get(prod_id, "")
        for variant in prod.get("variants", []):
            sku = str(variant.get("sku") or variant.get("id") or "")
            if not sku:
                continue
            sold = tn_lavan_sales.get(sku, 0)
            sold += tn_lavan_sales.get(str(variant.get("id", "")), 0)
            tn_lavan_sku_index[sku] = {
                "sold": sold,
                "1h":   tn_lavan_first.get(sku, 0),
                "2h":   tn_lavan_second.get(sku, 0),
                "url":  lavan_url,
            }

    # ── SKUs ya procesados (para evitar duplicados)
    seen_skus = set()
    rows = []

    # ── PASO 1: procesar todos los productos de TN Pret (fuente de stock)
    for prod in tn_pret_products:
        product_name     = get_name(prod)
        created_at       = prod.get("created_at", "")
        published        = prod.get("published", False)
        product_id       = str(prod.get("id", ""))
        tn_pret_url      = tn_pret_url_map.get(product_id, "")
        prod_days_active = days_active(created_at)
        is_new           = prod_days_active < DAYS

        for variant in prod.get("variants", []):
            sku        = str(variant.get("sku") or variant.get("id") or "")
            supplier   = detect_supplier(product_name, sku)
            stock      = variant.get("stock", 0) or 0
            price      = float(variant.get("price") or 0)
            cost       = float(variant.get("cost") or 0)
            var_name   = variant.get("values", [{}])
            var_label  = " / ".join(
                v.get("es") or v.get("en") or str(v) if isinstance(v, dict) else str(v)
                for v in var_name
            ) if var_name else ""
            stock_mgmt = variant.get("stock_management", True)

            seen_skus.add(sku)

            # Ventas TN Pret
            sold_tn_pret  = tn_pret_sales.get(sku, 0)
            sold_tn_pret += tn_pret_sales.get(str(variant.get("id", "")), 0)
            tn_p_1h = tn_pret_first.get(sku, 0)
            tn_p_2h = tn_pret_second.get(sku, 0)

            # Ventas TN Lavan (mismo SKU, stock compartido)
            lavan_idx     = tn_lavan_sku_index.get(sku, {})
            sold_tn_lavan = lavan_idx.get("sold", 0)
            tn_l_1h       = lavan_idx.get("1h", 0)
            tn_l_2h       = lavan_idx.get("2h", 0)
            url_tn_lavan  = lavan_idx.get("url", "")

            # Ventas ML (ambas cuentas)
            sold_ml_pret  = ml_pret_by_sku.get(sku, 0)
            sold_ml_lavan = ml_lavan_by_sku.get(sku, 0)
            ml_1h = ml_pret_1h_sku.get(sku, 0) + ml_lavan_1h_sku.get(sku, 0)
            ml_2h = ml_pret_2h_sku.get(sku, 0) + ml_lavan_2h_sku.get(sku, 0)

            total_sold = sold_tn_pret + sold_tn_lavan + sold_ml_pret + sold_ml_lavan
            total_1h   = tn_p_1h + tn_l_1h + ml_1h
            total_2h   = tn_p_2h + tn_l_2h + ml_2h

            active_days = prod_days_active
            vel_diaria  = round(total_sold / active_days, 4) if active_days > 0 else 0
            vel_semanal = round(vel_diaria * 7, 2)

            trend_label, trend_pct = calc_trend(total_1h, total_2h, active_days)

            if vel_diaria > 0 and stock_mgmt:
                dias_quiebre  = int(stock / vel_diaria)
                fecha_quiebre = (NOW + timedelta(days=dias_quiebre)).strftime("%d/%m/%Y")
            else:
                dias_quiebre  = 9999
                fecha_quiebre = "—"

            if not stock_mgmt:
                alerta = "⚪ Sin control"
            elif total_sold == 0:
                alerta = "⚫ Sin ventas"
            elif stock == 0 and vel_semanal >= 1:
                alerta = "🔴 SIN STOCK"
            elif dias_quiebre < 14:
                alerta = "🔴 CRÍTICO"
            elif dias_quiebre < 30:
                alerta = "🟡 BAJO"
            else:
                alerta = "🟢 OK"

            if active_days > 45:
                confianza = "🟢 Alta"
            elif active_days > 20:
                confianza = "🟡 Media"
            else:
                confianza = "🔴 Baja"

            margin_pct  = round((price - cost) / price * 100, 1) if price > 0 else 0
            margin_unit = round(price - cost, 2)
            revenue_60  = round(total_sold * price, 2)
            ganancia_60 = round(total_sold * margin_unit, 2)
            a_reponer   = units_to_order(stock, vel_diaria) if vel_diaria > 0 else 0

            canal_vals = {
                "ML Pret": sold_ml_pret, "ML Lavan": sold_ml_lavan,
                "TN Pret": sold_tn_pret, "TN Lavan": sold_tn_lavan,
            }
            canal_dom = max(canal_vals, key=canal_vals.get) if total_sold > 0 else "—"

            rows.append({
                "SKU":                    sku,
                "Producto":               product_name,
                "Variante":               var_label,
                "Marca":                  "Pret a Home / Casa Lavan",
                "Proveedor":              supplier,
                "Precio TN ($)":          price,
                "Costo ($)":              cost,
                "Margen ($)":             margin_unit,
                "Margen (%)":             margin_pct,
                "Publicado":              "Sí" if published else "No",
                "Fecha creación":         created_at[:10] if created_at else "",
                "Días activo (período)":  active_days,
                "¿Producto nuevo?":       "Sí" if is_new else "No",
                "Confianza métrica":      confianza,
                "Unid. ML Pret":          sold_ml_pret,
                "Unid. ML Lavan":         sold_ml_lavan,
                "Unid. TN Pret":          sold_tn_pret,
                "Unid. TN Lavan":         sold_tn_lavan,
                "Total vendido":          total_sold,
                "Vel. diaria":            vel_diaria,
                "Vel. semanal":           vel_semanal,
                "Tendencia":              trend_label,
                "Tendencia (%)":          trend_pct,
                "Stock actual":           stock if stock_mgmt else "Sin ctrl",
                "Días para quiebre":      dias_quiebre if dias_quiebre < 9999 else "—",
                "Fecha quiebre est.":     fecha_quiebre,
                "Alerta stock":           alerta,
                "A reponer (uds)":        a_reponer if vel_diaria > 0 else "—",
                "Canal dominante":        canal_dom,
                "Revenue 60d ($)":        revenue_60,
                "Ganancia bruta 60d ($)": ganancia_60,
                "URL TN Pret":            tn_pret_url,
                "URL TN Lavan":           url_tn_lavan,
                "URL ML Pret":            ml_pret_sku_to_url.get(sku, ""),
                "URL ML Lavan":           ml_lavan_sku_to_url.get(sku, ""),
            })

    # ── PASO 2: SKUs exclusivos de TN Lavan (no están en TN Pret)
    for prod in tn_lavan_products:
        product_name     = get_name(prod)
        created_at       = prod.get("created_at", "")
        published        = prod.get("published", False)
        product_id       = str(prod.get("id", ""))
        url_tn_lavan     = tn_lavan_url_map.get(product_id, "")
        prod_days_active = days_active(created_at)
        is_new           = prod_days_active < DAYS

        for variant in prod.get("variants", []):
            sku = str(variant.get("sku") or variant.get("id") or "")
            if sku in seen_skus:
                continue  # Ya procesado desde TN Pret
            seen_skus.add(sku)
            supplier = detect_supplier(product_name, sku)

            stock      = variant.get("stock", 0) or 0
            price      = float(variant.get("price") or 0)
            cost       = float(variant.get("cost") or 0)
            var_name   = variant.get("values", [{}])
            var_label  = " / ".join(
                v.get("es") or v.get("en") or str(v) if isinstance(v, dict) else str(v)
                for v in var_name
            ) if var_name else ""
            stock_mgmt = variant.get("stock_management", True)

            sold_tn_lavan  = tn_lavan_sales.get(sku, 0)
            sold_tn_lavan += tn_lavan_sales.get(str(variant.get("id", "")), 0)
            tn_l_1h = tn_lavan_first.get(sku, 0)
            tn_l_2h = tn_lavan_second.get(sku, 0)

            sold_ml_pret  = ml_pret_by_sku.get(sku, 0)
            sold_ml_lavan = ml_lavan_by_sku.get(sku, 0)
            ml_1h = ml_pret_1h_sku.get(sku, 0) + ml_lavan_1h_sku.get(sku, 0)
            ml_2h = ml_pret_2h_sku.get(sku, 0) + ml_lavan_2h_sku.get(sku, 0)

            total_sold = sold_tn_lavan + sold_ml_pret + sold_ml_lavan
            total_1h   = tn_l_1h + ml_1h
            total_2h   = tn_l_2h + ml_2h

            active_days = prod_days_active
            vel_diaria  = round(total_sold / active_days, 4) if active_days > 0 else 0
            vel_semanal = round(vel_diaria * 7, 2)

            trend_label, trend_pct = calc_trend(total_1h, total_2h, active_days)

            if vel_diaria > 0 and stock_mgmt:
                dias_quiebre  = int(stock / vel_diaria)
                fecha_quiebre = (NOW + timedelta(days=dias_quiebre)).strftime("%d/%m/%Y")
            else:
                dias_quiebre  = 9999
                fecha_quiebre = "—"

            if not stock_mgmt:
                alerta = "⚪ Sin control"
            elif total_sold == 0:
                alerta = "⚫ Sin ventas"
            elif stock == 0 and vel_semanal >= 1:
                alerta = "🔴 SIN STOCK"
            elif dias_quiebre < 14:
                alerta = "🔴 CRÍTICO"
            elif dias_quiebre < 30:
                alerta = "🟡 BAJO"
            else:
                alerta = "🟢 OK"

            if active_days > 45:
                confianza = "🟢 Alta"
            elif active_days > 20:
                confianza = "🟡 Media"
            else:
                confianza = "🔴 Baja"

            margin_pct  = round((price - cost) / price * 100, 1) if price > 0 else 0
            margin_unit = round(price - cost, 2)
            revenue_60  = round(total_sold * price, 2)
            ganancia_60 = round(total_sold * margin_unit, 2)
            a_reponer   = units_to_order(stock, vel_diaria) if vel_diaria > 0 else 0

            canal_vals = {
                "ML Pret": sold_ml_pret, "ML Lavan": sold_ml_lavan,
                "TN Pret": 0, "TN Lavan": sold_tn_lavan,
            }
            canal_dom = max(canal_vals, key=canal_vals.get) if total_sold > 0 else "—"

            rows.append({
                "SKU":                    sku,
                "Producto":               product_name,
                "Variante":               var_label,
                "Marca":                  "Casa Lavan",
                "Proveedor":              supplier,
                "Precio TN ($)":          price,
                "Costo ($)":              cost,
                "Margen ($)":             margin_unit,
                "Margen (%)":             margin_pct,
                "Publicado":              "Sí" if published else "No",
                "Fecha creación":         created_at[:10] if created_at else "",
                "Días activo (período)":  active_days,
                "¿Producto nuevo?":       "Sí" if is_new else "No",
                "Confianza métrica":      confianza,
                "Unid. ML Pret":          sold_ml_pret,
                "Unid. ML Lavan":         sold_ml_lavan,
                "Unid. TN Pret":          0,
                "Unid. TN Lavan":         sold_tn_lavan,
                "Total vendido":          total_sold,
                "Vel. diaria":            vel_diaria,
                "Vel. semanal":           vel_semanal,
                "Tendencia":              trend_label,
                "Tendencia (%)":          trend_pct,
                "Stock actual":           stock if stock_mgmt else "Sin ctrl",
                "Días para quiebre":      dias_quiebre if dias_quiebre < 9999 else "—",
                "Fecha quiebre est.":     fecha_quiebre,
                "Alerta stock":           alerta,
                "A reponer (uds)":        a_reponer if vel_diaria > 0 else "—",
                "Canal dominante":        canal_dom,
                "Revenue 60d ($)":        revenue_60,
                "Ganancia bruta 60d ($)": ganancia_60,
                "URL TN Pret":            "",
                "URL TN Lavan":           url_tn_lavan,
                "URL ML Pret":            ml_pret_sku_to_url.get(sku, ""),
                "URL ML Lavan":           ml_lavan_sku_to_url.get(sku, ""),
            })

    return rows

# ─── EXCEL ────────────────────────────────────────────────────────────────────


# ─── SOBRESTOCK SCORING ────────────────────────────────────────────────────────

DEAD_STOCK_DAYS   = 60    # sin ventas en X días → stock muerto
SLOW_VEL_WEEKLY   = 0.5   # vel semanal < X → lento
SLOW_DIAS_QUIEBRE = 90    # días quiebre > X → lento
OVER_DIAS_QUIEBRE = 60    # días quiebre > X → sobrestock

def capital_inmovilizado(row):
    stock = row.get("Stock actual", 0)
    cost  = row.get("Costo ($)", 0)
    if isinstance(stock, int) and isinstance(cost, (int, float)) and cost > 0:
        return round(stock * cost, 2)
    return 0

def sobrestock_category(row):
    """Returns (category_label, accion) or None if not overstock."""
    stock     = row.get("Stock actual", 0)
    vel_sem   = row.get("Vel. semanal", 0)
    total_sol = row.get("Total vendido", 0)
    dias_q    = row.get("Días para quiebre", 9999)
    stock_mgmt = isinstance(stock, int)  # "Sin ctrl" is str

    if not stock_mgmt or stock == 0:
        return None

    no_ventas  = (total_sol == 0)
    dias_q_int = dias_q if isinstance(dias_q, int) else 9999

    if no_ventas:
        return ("🔴 Stock muerto", "Liquidar / Devolver")
    elif vel_sem < SLOW_VEL_WEEKLY and dias_q_int > SLOW_DIAS_QUIEBRE:
        return ("🟠 Stock lento", "Promocionar activamente")
    elif dias_q_int > OVER_DIAS_QUIEBRE:
        return ("🟡 Sobrestock", "Monitorear / Descuento leve")
    return None

# ─── EXCEL HELPERS (shared) ───────────────────────────────────────────────────

def make_wb_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    CLR = {
        "HEADER_BG":   "1F3864", "HEADER_FG":   "FFFFFF",
        "SECTION_BG":  "D6E4F0", "ROW_ALT":     "F2F7FB",
        "CRITICAL":    "FFCCCC", "NO_STOCK":    "FF9999",
        "LOW":         "FFF2CC", "OK":          "CCFFCC",
        "NO_SALES":    "EEEEEE", "LOW_CONF":    "FFE6CC",
        "PROVEEDOR":   "2E4057", "ACCEL":       "C6EFCE",
        "DECEL":       "FFCCCC", "DEAD":        "FF6B6B",
        "SLOW":        "FFB347", "OVER":        "FFE066",
    }
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(color):
        return PatternFill("solid", start_color=color, fgColor=color)
    def hfont(bold=False, color="000000", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    return CLR, border, hfill, hfont, center

def write_url_cell_fn(cell, url, hfont_fn, label="Ver"):
    from openpyxl.styles import Font, Alignment
    if url:
        cell.value     = label
        cell.hyperlink = url
        cell.font      = Font(name="Arial", size=10, color="0563C1", underline="single")
    else:
        cell.value = "—"
        cell.font  = hfont_fn(color="999999")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def write_header_row_fn(ws, headers_widths, CLR, border, hfill, hfont, center, row=2, bg=None):
    from openpyxl.utils import get_column_letter
    bg = bg or CLR["HEADER_BG"]
    for ci, (h, w) in enumerate(headers_widths, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = hfont(bold=True, color="FFFFFF")
        c.fill      = hfill(bg)
        c.alignment = center()
        c.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 18

def write_sheet_title_fn(ws, title, ncols, CLR, hfill, hfont, center, bg=None):
    from openpyxl.utils import get_column_letter
    bg = bg or CLR["HEADER_BG"]
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"]           = title
    ws["A1"].font      = hfont(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill      = hfill(bg)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 22

# ─── HOJA VENTAS COMPLETA (shared entre ambos excels) ─────────────────────────

def write_ventas_sheet(wb, rows, CLR, border, hfill, hfont, center):
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    ws = wb.active
    ws.title = f"Ventas {DAYS}d"
    ws.freeze_panes = "A3"

    COLS = [
        ("SKU",           14), ("Producto",       34), ("Variante",    18),
        ("Marca",         14), ("Proveedor",       16),
        ("Precio ($)",    13), ("Costo ($)",       11), ("Margen ($)",  11), ("Margen (%)", 10),
        ("Publicado",     10), ("Fecha creación",  14), ("Días activo", 12),
        ("¿Nuevo?",        9), ("Confianza",       12),
        ("ML Pret",       13), ("ML Lavan",        13), ("TN Pret",     13), ("TN Lavan",    13),
        ("Total vendido", 13), ("Vel. diaria",     11), ("Vel. semanal",11),
        ("Tendencia",     14), ("Tend. (%)",       10),
        ("Stock",         12), ("Días quiebre",    12), ("Fecha quiebre",14),
        ("Alerta",        16), ("A reponer",       12), ("Canal dom.",  14),
        ("Revenue ($)",   15), ("Ganancia ($)",    15),
        ("TN Pret URL",   10), ("TN Lavan URL",    10),
        ("ML Pret URL",   10), ("ML Lavan URL",    10),
    ]
    col_keys = [
        "SKU", "Producto", "Variante", "Marca", "Proveedor",
        "Precio TN ($)", "Costo ($)", "Margen ($)", "Margen (%)",
        "Publicado", "Fecha creación", "Días activo (período)", "¿Producto nuevo?",
        "Confianza métrica",
        "Unid. ML Pret", "Unid. ML Lavan", "Unid. TN Pret", "Unid. TN Lavan",
        "Total vendido", "Vel. diaria", "Vel. semanal",
        "Tendencia", "Tendencia (%)",
        "Stock actual", "Días para quiebre", "Fecha quiebre est.",
        "Alerta stock", "A reponer (uds)", "Canal dominante",
        "Revenue 60d ($)", "Ganancia bruta 60d ($)",
        "URL TN Pret", "URL TN Lavan", "URL ML Pret", "URL ML Lavan",
    ]
    url_keys = {"URL TN Pret", "URL TN Lavan", "URL ML Pret", "URL ML Lavan"}

    write_sheet_title_fn(ws,
        f"Reporte de Ventas — Últimos {DAYS} días  |  "
        f"{DATE_FROM_STR} → {DATE_TO_STR}  |  "
        f"Generado: {NOW.strftime('%d/%m/%Y %H:%M')} UTC",
        len(COLS), CLR, hfill, hfont, center)
    write_header_row_fn(ws, COLS, CLR, border, hfill, hfont, center, row=2)

    for ri, row in enumerate(rows, 3):
        alerta = row.get("Alerta stock", "")
        confia = row.get("Confianza métrica", "")
        is_alt = (ri % 2 == 0)

        if "SIN STOCK" in alerta:    row_bg = CLR["NO_STOCK"]
        elif "CRÍTICO" in alerta:    row_bg = CLR["CRITICAL"]
        elif "BAJO"    in alerta:    row_bg = CLR["LOW"]
        elif "Sin ventas" in alerta: row_bg = CLR["NO_SALES"]
        elif "Baja" in confia:       row_bg = CLR["LOW_CONF"]
        elif is_alt:                 row_bg = CLR["ROW_ALT"]
        else:                        row_bg = "FFFFFF"

        for ci, key in enumerate(col_keys, 1):
            val  = row.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font   = hfont()
            cell.fill   = hfill(row_bg)
            cell.border = border

            if key in url_keys:
                write_url_cell_fn(cell, val, hfont)
                continue
            if key in ("Precio TN ($)", "Costo ($)", "Margen ($)",
                       "Revenue 60d ($)", "Ganancia bruta 60d ($)"):
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            elif key == "Margen (%)":
                if isinstance(val, (int, float)):
                    cell.number_format = '0.0"%"'
                    cell.alignment = Alignment(horizontal="right")
            elif key in ("Vel. diaria", "Vel. semanal"):
                if isinstance(val, (int, float)):
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right")
            elif key == "Tendencia (%)":
                if isinstance(val, (int, float)):
                    cell.number_format = '+0.0;-0.0;0.0'
                    cell.alignment = Alignment(horizontal="right")
                    if val >= 20:   cell.fill = hfill(CLR["ACCEL"])
                    elif val <= -20: cell.fill = hfill(CLR["DECEL"])
            elif key in ("Unid. ML Pret", "Unid. ML Lavan", "Unid. TN Pret", "Unid. TN Lavan",
                         "Total vendido", "Stock actual", "Días para quiebre",
                         "Días activo (período)", "A reponer (uds)"):
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="center")
            elif key == "Producto":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(
                    horizontal="center" if key in (
                        "Alerta stock", "Confianza métrica", "Canal dominante",
                        "¿Producto nuevo?", "Publicado", "Tendencia", "Proveedor",
                    ) else "left", vertical="center")
        ws.row_dimensions[ri].height = 15

    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{2 + len(rows)}"

# ─── WRITE EXCEL REPOSICIÓN ───────────────────────────────────────────────────

def write_excel_reposicion(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    CLR, border, hfill, hfont, center = make_wb_styles()
    wb = Workbook()

    # HOJA 1 — Ventas completa
    write_ventas_sheet(wb, rows, CLR, border, hfill, hfont, center)

    # HOJA 2 — Resumen Canales
    ws2 = wb.create_sheet("Resumen Canales")
    canal_headers = [
        ("Canal", 20), ("Órdenes", 12), ("Unidades", 18),
        ("Revenue ($)", 16), ("Ticket prom. ($)", 16), ("Top SKU", 22),
    ]
    write_sheet_title_fn(ws2, f"Resumen por Canal — Últimos {DAYS} días",
                         len(canal_headers), CLR, hfill, hfont, center)
    write_header_row_fn(ws2, canal_headers, CLR, border, hfill, hfont, center, row=2)

    channel_summaries = []
    for key, cfg in channels.items():
        res = results.get(key, {})
        orders = res.get("orders", [])
        n_orders = len(orders)
        if cfg["type"] == "tiendanube":
            units = sum(res.get("sales", {}).values())
            rev   = sum(
                float(item.get("price") or 0) * (item.get("quantity") or 1)
                for o in orders for item in o.get("products", [])
            )
            top_sku = max(res.get("sales", {}).items(), key=lambda x: x[1])[0] if res.get("sales") else "—"
        else:
            sales = res.get("sales", defaultdict(int))
            sku_totals = defaultdict(int)
            for (kind, k), v in sales.items():
                sku_totals[k] += v
            units   = sum(sku_totals.values())
            rev     = sum(float(o.get("total_amount", 0)) for o in orders)
            top_sku = max(sku_totals.items(), key=lambda x: x[1])[0] if sku_totals else "—"
        ticket = round(rev / n_orders, 2) if n_orders > 0 else 0
        channel_summaries.append((cfg["label"], n_orders, units, rev, ticket, top_sku))

    for ri, row_data in enumerate(channel_summaries, 3):
        bg = CLR["ROW_ALT"] if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = hfont(); c.fill = hfill(bg); c.border = border
            if ci in (2, 3):
                c.number_format = '#,##0'; c.alignment = Alignment(horizontal="center")
            elif ci in (4, 5):
                c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    ri_tot = 3 + len(channel_summaries)
    for ci in range(1, 7):
        c = ws2.cell(ri_tot, ci); c.font = hfont(bold=True)
        c.fill = hfill(CLR["SECTION_BG"]); c.border = border
    ws2.cell(ri_tot, 1).value = "TOTAL"
    ws2.cell(ri_tot, 2).value = f"=SUM(B3:B{ri_tot-1})"; ws2.cell(ri_tot, 2).number_format = '#,##0'
    ws2.cell(ri_tot, 3).value = f"=SUM(C3:C{ri_tot-1})"; ws2.cell(ri_tot, 3).number_format = '#,##0'
    ws2.cell(ri_tot, 4).value = f"=SUM(D3:D{ri_tot-1})"; ws2.cell(ri_tot, 4).number_format = '#,##0.00'

    # Columnas comunes para pestañas de stock
    stock_cols = [
        ("Alerta",        14), ("SKU",          14), ("Producto",     32), ("Variante",    16),
        ("Proveedor",     16), ("Marca",         14), ("Stock",        12), ("Vel. semanal",12),
        ("Tendencia",     14), ("Días quiebre",  13), ("Fecha quiebre",14), ("A reponer",   12),
        ("Confianza",     12), ("Capital inmov.",15),
        ("TN Pret",       10), ("TN Lavan",      10), ("ML Pret",      10), ("ML Lavan",    10),
    ]
    stock_keys = [
        "Alerta stock", "SKU", "Producto", "Variante",
        "Proveedor", "Marca", "Stock actual", "Vel. semanal",
        "Tendencia", "Días para quiebre", "Fecha quiebre est.", "A reponer (uds)",
        "Confianza métrica", "Capital inmovilizado ($)",
        "URL TN Pret", "URL TN Lavan", "URL ML Pret", "URL ML Lavan",
    ]
    url_cols_set = {"URL TN Pret", "URL TN Lavan", "URL ML Pret", "URL ML Lavan"}

    def write_stock_sheet(ws, title, sheet_rows, hdr_bg, sort_key=None):
        write_sheet_title_fn(ws, title, len(stock_cols), CLR, hfill, hfont, center, bg=hdr_bg)
        write_header_row_fn(ws, stock_cols, CLR, border, hfill, hfont, center, row=2, bg=hdr_bg)
        if sort_key:
            sheet_rows = sorted(sheet_rows, key=sort_key)
        for ri, row in enumerate(sheet_rows, 3):
            alerta = row.get("Alerta stock", "")
            is_alt = (ri % 2 == 0)
            if "SIN STOCK" in alerta:    bg = CLR["NO_STOCK"]
            elif "CRÍTICO" in alerta:    bg = CLR["CRITICAL"]
            elif "BAJO"    in alerta:    bg = CLR["LOW"]
            elif is_alt:                 bg = CLR["ROW_ALT"]
            else:                        bg = "FFFFFF"
            for ci, key in enumerate(stock_keys, 1):
                val = row.get(key, "")
                c   = ws.cell(row=ri, column=ci, value=val)
                c.font = hfont(); c.fill = hfill(bg); c.border = border
                if key in url_cols_set:
                    write_url_cell_fn(c, val, hfont); continue
                if key == "Vel. semanal" and isinstance(val, float):
                    c.number_format = '0.00'; c.alignment = Alignment(horizontal="center")
                elif key == "Capital inmovilizado ($)" and isinstance(val, (int, float)):
                    c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
                elif key in ("Stock actual", "Días para quiebre", "A reponer (uds)") and isinstance(val, int):
                    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="center")
                elif key == "Tendencia":
                    tend_pct = row.get("Tendencia (%)")
                    if isinstance(tend_pct, (int, float)):
                        if tend_pct >= 20:   c.fill = hfill(CLR["ACCEL"])
                        elif tend_pct <= -20: c.fill = hfill(CLR["DECEL"])
                    c.alignment = Alignment(horizontal="center")
                else:
                    c.alignment = Alignment(
                        horizontal="left" if key in ("Producto", "Variante", "SKU") else "center",
                        vertical="center")
            ws.row_dimensions[ri].height = 15
        ws.auto_filter.ref = f"A2:{get_column_letter(len(stock_cols))}{2+len(sheet_rows)}"

    # Enriquecer rows con capital inmovilizado
    for r in rows:
        r["Capital inmovilizado ($)"] = capital_inmovilizado(r)

    # HOJA 3 — Stock quebrado CON ventas (últimos 60d)
    ws3 = wb.create_sheet("🔴 Quebrado c-ventas")
    quebrado_con_ventas = [
        r for r in rows
        if isinstance(r.get("Stock actual"), int)
        and r.get("Stock actual", 0) == 0
        and r.get("Total vendido", 0) > 0
    ]
    write_stock_sheet(ws3,
        f"🔴 Stock Quebrado CON Ventas últimos {DAYS}d  |  {DATE_TO_STR}",
        quebrado_con_ventas, "C00000",
        sort_key=lambda r: -r.get("Vel. semanal", 0))

    # HOJA 4 — Stock quebrado SIN ventas
    ws4 = wb.create_sheet("⚫ Quebrado s-ventas")
    quebrado_sin_ventas = [
        r for r in rows
        if isinstance(r.get("Stock actual"), int)
        and r.get("Stock actual", 0) == 0
        and r.get("Total vendido", 0) == 0
    ]
    write_stock_sheet(ws4,
        f"⚫ Stock Quebrado SIN Ventas  |  {DATE_TO_STR}",
        quebrado_sin_ventas, "555555")

    # HOJA 5 — Por quebrar en menos de 60 días
    ws5 = wb.create_sheet("⚠ Por quebrar <60d")
    por_quebrar = [
        r for r in rows
        if isinstance(r.get("Stock actual"), int)
        and r.get("Stock actual", 0) > 0
        and isinstance(r.get("Días para quiebre"), int)
        and r.get("Días para quiebre", 9999) < 60
        and r.get("Vel. semanal", 0) >= 0.5
    ]
    write_stock_sheet(ws5,
        f"⚠ Por Quebrar en <60 días según vel. de venta  |  {DATE_TO_STR}",
        por_quebrar, "E07B00",
        sort_key=lambda r: r.get("Días para quiebre", 9999) if isinstance(r.get("Días para quiebre"), int) else 9999)

    # HOJAS por Proveedor — una pestaña por cada proveedor
    # Columnas exactas pedidas: SKU, Producto, Variante, Proveedor, Stock,
    #                            Vel.Semanal, Días quiebre, Fecha quiebre, A reponer
    prov_sheet_cols = [
        ("SKU",            14), ("Producto",      36), ("Variante",     16),
        ("Proveedor",      16), ("Stock",         10), ("Vel. semanal", 13),
        ("Días quiebre",   13), ("Fecha quiebre", 14), ("A reponer",    13),
    ]
    prov_sheet_keys = [
        "SKU", "Producto", "Variante",
        "Proveedor", "Stock actual", "Vel. semanal",
        "Días para quiebre", "Fecha quiebre est.", "A reponer (uds)",
    ]
    url_cols_prov = set()  # sin columnas de links en pestañas de proveedor

    # Paleta de colores por proveedor (bg del header)
    PROV_COLORS = [
        "1A237E","1B5E20","B71C1C","E65100","4A148C",
        "006064","33691E","880E4F","0D47A1","BF360C",
        "37474F","1565C0","2E7D32","6A1B9A","00695C","F57F17",
    ]

    # Agrupar SKUs por proveedor — SOLO los que hay que reponer:
    #   1) Stock = 0 y tuvo ventas en los últimos 60d (quebrado con demanda)
    #   2) Stock > 0 pero se queda sin stock en los próximos 10 días
    prov_groups = defaultdict(list)
    for r in rows:
        stock   = r.get("Stock actual", 0)
        total_v = r.get("Total vendido", 0)
        dias_q  = r.get("Días para quiebre", 9999)
        vel_d   = r.get("Vel. diaria", 0)

        if not isinstance(stock, int):   # "Sin ctrl" → saltar
            continue

        es_quebrado_con_ventas = (stock == 0 and total_v > 0)
        dias_q_int = dias_q if isinstance(dias_q, int) else 9999
        es_por_quebrar = (stock >= 0 and dias_q_int <= 10 and vel_d > 0 and total_v > 0)

        if es_quebrado_con_ventas or es_por_quebrar:
            prov = (r.get("Proveedor") or SUPPLIER_OTHER).strip().title()
            stk   = stock
            a_rep = max(0, round(vel_d * COVERAGE_DAYS - stk)) if vel_d > 0 else 0
            row_copy = dict(r)
            row_copy["A reponer (uds)"] = a_rep
            prov_groups[prov].append(row_copy)

    known_order = [s for s, _ in SUPPLIER_KEYWORDS] + [SUPPLIER_OTHER]
    extra_provs = sorted(p for p in prov_groups if p not in known_order)
    ordered_provs = [s for s in known_order if s in prov_groups] + extra_provs

    for pi, prov in enumerate(ordered_provs):
        group = sorted(prov_groups[prov],
                       key=lambda r: r.get("A reponer (uds)", 0) if isinstance(r.get("A reponer (uds)"), (int,float)) else 0,
                       reverse=True)
        # Sheet name: max 31 chars, no special chars
        sheet_name = prov[:28] if len(prov) > 28 else prov
        ws_p = wb.create_sheet(f"📦 {sheet_name}")
        hdr_color = PROV_COLORS[pi % len(PROV_COLORS)]

        write_sheet_title_fn(ws_p,
            f"📦 {prov}  —  {len(group)} SKUs a reponer (sin stock c/ventas + quiebre ≤10d)  |  {DATE_TO_STR}",
            len(prov_sheet_cols), CLR, hfill, hfont, center, bg=hdr_color)
        write_header_row_fn(ws_p, prov_sheet_cols, CLR, border, hfill, hfont, center, row=2, bg=hdr_color)

        for ri, row in enumerate(group, 3):
            stock_val = row.get("Stock actual", 0)
            dias_q    = row.get("Días para quiebre", 9999)
            vel_s     = row.get("Vel. semanal", 0)
            is_alt    = (ri % 2 == 0)

            if isinstance(stock_val, int) and stock_val == 0 and vel_s > 0:
                bg = CLR["NO_STOCK"]
            elif isinstance(dias_q, int) and dias_q < 14:
                bg = CLR["CRITICAL"]
            elif isinstance(dias_q, int) and dias_q < 30:
                bg = CLR["LOW"]
            elif is_alt:
                bg = CLR["ROW_ALT"]
            else:
                bg = "FFFFFF"

            from openpyxl.styles import Alignment
            for ci, key in enumerate(prov_sheet_keys, 1):
                val = row.get(key, "")
                c   = ws_p.cell(row=ri, column=ci, value=val)
                c.font = hfont(); c.fill = hfill(bg); c.border = border
                if key in url_cols_prov:
                    write_url_cell_fn(c, val, hfont)
                    continue
                if key == "Vel. semanal" and isinstance(val, float):
                    c.number_format = "0.00"; c.alignment = Alignment(horizontal="center")
                elif key in ("Stock actual", "Días para quiebre", "A reponer (uds)") and isinstance(val, int):
                    c.number_format = "#,##0"; c.alignment = Alignment(horizontal="center")
                elif key == "Producto":
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    c.alignment = Alignment(horizontal="left" if key in ("SKU","Variante") else "center",
                                            vertical="center")
            ws_p.row_dimensions[ri].height = 15

        from openpyxl.utils import get_column_letter
        ws_p.auto_filter.ref = f"A2:{get_column_letter(len(prov_sheet_cols))}{2+len(group)}"

        # Fila resumen al final
        ri_sum = 3 + len(group)
        total_rep = sum(r.get("A reponer (uds)", 0) for r in group if isinstance(r.get("A reponer (uds)"), int))
        ws_p.merge_cells(f"A{ri_sum}:{get_column_letter(len(prov_sheet_cols))}{ri_sum}")
        c_sum = ws_p.cell(ri_sum, 1,
                          value=f"Total a reponer (10d cobertura): {total_rep} uds  |  {len(group)} SKUs a reponer")
        c_sum.font = hfont(bold=True, color="FFFFFF")
        c_sum.fill = hfill(hdr_color)
        c_sum.alignment = Alignment(horizontal="left", vertical="center")
        ws_p.row_dimensions[ri_sum].height = 16


    wb.save(path)
    total_prov_skus = sum(len(v) for v in prov_groups.values())
    print(f"\n  ✅ Reposición guardado: {path}")
    print(f"     {len(rows)} SKUs totales (sin duplicados)")
    print(f"     {len(quebrado_con_ventas)} con stock quebrado y demanda")
    print(f"     {len(quebrado_sin_ventas)} con stock quebrado sin demanda")
    print(f"     {len(por_quebrar)} por quebrar en <60d")
    print(f"     {len(ordered_provs)} pestañas de proveedor | {total_prov_skus} SKUs a reponer")


# ─── WRITE EXCEL SOBRESTOCK ───────────────────────────────────────────────────

def write_excel_sobrestock(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from collections import OrderedDict

    CLR, border, hfill, hfont, center = make_wb_styles()
    wb = Workbook()

    # HOJA 1 — Ventas completa
    write_ventas_sheet(wb, rows, CLR, border, hfill, hfont, center)

    # Enriquecer con categoría sobrestock y capital inmovilizado
    for r in rows:
        r["Capital inmovilizado ($)"] = capital_inmovilizado(r)
        cat = sobrestock_category(r)
        r["Cat. sobrestock"]  = cat[0] if cat else ""
        r["Acción sugerida"]  = cat[1] if cat else ""

    over_cols = [
        ("Categoría",      18), ("Acción sugerida", 22),
        ("SKU",            14), ("Producto",         34), ("Variante",    16),
        ("Proveedor",      16), ("Marca",            14),
        ("Stock actual",   12), ("Vel. semanal",     12), ("Tendencia",   14),
        ("Días quiebre",   13), ("Capital inmov. ($)",16),
        ("Precio ($)",     13), ("Costo ($)",         11),
        ("TN Pret",        10), ("TN Lavan",          10),
        ("ML Pret",        10), ("ML Lavan",          10),
    ]
    over_keys = [
        "Cat. sobrestock", "Acción sugerida",
        "SKU", "Producto", "Variante",
        "Proveedor", "Marca",
        "Stock actual", "Vel. semanal", "Tendencia",
        "Días para quiebre", "Capital inmovilizado ($)",
        "Precio TN ($)", "Costo ($)",
        "URL TN Pret", "URL TN Lavan", "URL ML Pret", "URL ML Lavan",
    ]
    url_cols_set = {"URL TN Pret", "URL TN Lavan", "URL ML Pret", "URL ML Lavan"}

    def write_over_sheet(ws, title, sheet_rows, hdr_bg):
        write_sheet_title_fn(ws, title, len(over_cols), CLR, hfill, hfont, center, bg=hdr_bg)
        write_header_row_fn(ws, over_cols, CLR, border, hfill, hfont, center, row=2, bg=hdr_bg)
        # Ordenar por capital inmovilizado desc
        sheet_rows = sorted(sheet_rows,
                            key=lambda r: r.get("Capital inmovilizado ($)", 0), reverse=True)
        for ri, row in enumerate(sheet_rows, 3):
            cat    = row.get("Cat. sobrestock", "")
            is_alt = (ri % 2 == 0)
            if   "muerto" in cat: bg = "FFCDD2"
            elif "lento"  in cat: bg = "FFE0B2"
            elif "sobre"  in cat.lower(): bg = "FFF9C4"
            elif is_alt:          bg = CLR["ROW_ALT"]
            else:                 bg = "FFFFFF"

            for ci, key in enumerate(over_keys, 1):
                val = row.get(key, "")
                c   = ws.cell(row=ri, column=ci, value=val)
                c.font = hfont(); c.fill = hfill(bg); c.border = border
                if key in url_cols_set:
                    write_url_cell_fn(c, val, hfont); continue
                if key in ("Capital inmovilizado ($)", "Precio TN ($)", "Costo ($)") and isinstance(val, (int, float)):
                    c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
                elif key == "Vel. semanal" and isinstance(val, float):
                    c.number_format = '0.00'; c.alignment = Alignment(horizontal="center")
                elif key in ("Stock actual", "Días para quiebre") and isinstance(val, int):
                    c.number_format = '#,##0'; c.alignment = Alignment(horizontal="center")
                elif key == "Tendencia":
                    tend_pct = row.get("Tendencia (%)")
                    if isinstance(tend_pct, (int, float)):
                        if tend_pct >= 20:    c.fill = hfill(CLR["ACCEL"])
                        elif tend_pct <= -20: c.fill = hfill(CLR["DECEL"])
                    c.alignment = Alignment(horizontal="center")
                elif key in ("Producto", "Variante", "SKU", "Acción sugerida"):
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[ri].height = 15
        ws.auto_filter.ref = f"A2:{get_column_letter(len(over_cols))}{2+len(sheet_rows)}"

    # Separar categorías
    dead_rows = [r for r in rows if r.get("Cat. sobrestock") == "🔴 Stock muerto"]
    slow_rows = [r for r in rows if r.get("Cat. sobrestock") == "🟠 Stock lento"]
    over_rows = [r for r in rows if r.get("Cat. sobrestock") == "🟡 Sobrestock"]

    # HOJA 2 — Stock Muerto
    ws2 = wb.create_sheet("🔴 Stock Muerto")
    write_over_sheet(ws2,
        f"🔴 Stock Muerto — Sin ventas en {DAYS}d con stock > 0  |  {DATE_TO_STR}",
        dead_rows, "B71C1C")

    # HOJA 3 — Stock Lento
    ws3 = wb.create_sheet("🟠 Stock Lento")
    write_over_sheet(ws3,
        f"🟠 Stock Lento — Vel. <{SLOW_VEL_WEEKLY} u/sem y quiebre >{SLOW_DIAS_QUIEBRE}d  |  {DATE_TO_STR}",
        slow_rows, "E65100")

    # HOJA 4 — Sobrestock
    ws4 = wb.create_sheet("🟡 Sobrestock")
    write_over_sheet(ws4,
        f"🟡 Sobrestock — Quiebre en >{OVER_DIAS_QUIEBRE}d con algo de demanda  |  {DATE_TO_STR}",
        over_rows, "F57F17")

    # HOJA 5 — Resumen por Proveedor
    ws5 = wb.create_sheet("📊 Resumen por Proveedor")
    all_over = dead_rows + slow_rows + over_rows

    prov_data = defaultdict(lambda: {
        "muerto": 0, "lento": 0, "sobre": 0,
        "capital": 0.0, "skus": set()
    })
    for r in all_over:
        prov  = r.get("Proveedor", SUPPLIER_OTHER)
        cat   = r.get("Cat. sobrestock", "")
        cap   = r.get("Capital inmovilizado ($)", 0)
        sku   = r.get("SKU", "")
        prov_data[prov]["capital"] += cap
        prov_data[prov]["skus"].add(sku)
        if "muerto" in cat:        prov_data[prov]["muerto"] += 1
        elif "lento"  in cat:      prov_data[prov]["lento"]  += 1
        elif "sobre" in cat.lower(): prov_data[prov]["sobre"] += 1

    res_cols = [
        ("Proveedor", 20), ("SKUs problemáticos", 18),
        ("🔴 Muertos", 14), ("🟠 Lentos", 14), ("🟡 Sobrestock", 14),
        ("Capital inmovilizado ($)", 22), ("Acción prioritaria", 24),
    ]
    write_sheet_title_fn(ws5,
        f"📊 Resumen Sobrestock por Proveedor  |  {DATE_TO_STR}",
        len(res_cols), CLR, hfill, hfont, center, bg="37474F")
    write_header_row_fn(ws5, res_cols, CLR, border, hfill, hfont, center, row=2, bg="37474F")

    prov_rows_sorted = sorted(prov_data.items(), key=lambda x: -x[1]["capital"])
    for ri, (prov, d) in enumerate(prov_rows_sorted, 3):
        is_alt = (ri % 2 == 0)
        bg = CLR["ROW_ALT"] if is_alt else "FFFFFF"
        total_skus = len(d["skus"])
        capital    = d["capital"]
        if d["muerto"] > 0:         accion = "🔴 Liquidar stock muerto"
        elif d["lento"] > total_skus // 2: accion = "🟠 Plan de promoción urgente"
        else:                        accion = "🟡 Revisar precios / descuentos"

        vals = [prov, total_skus, d["muerto"], d["lento"], d["sobre"], capital, accion]
        for ci, val in enumerate(vals, 1):
            c = ws5.cell(row=ri, column=ci, value=val)
            c.font = hfont(); c.fill = hfill(bg); c.border = border
            if ci == 6:
                c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
            elif ci in (2, 3, 4, 5):
                c.number_format = '#,##0'; c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws5.row_dimensions[ri].height = 15

    # Fila total
    ri_t = 3 + len(prov_rows_sorted)
    total_capital = sum(d["capital"] for d in prov_data.values())
    total_skus_all = len(set(r.get("SKU","") for r in all_over))
    for ci in range(1, 8):
        c = ws5.cell(ri_t, ci); c.font = hfont(bold=True)
        c.fill = hfill(CLR["SECTION_BG"]); c.border = border
    ws5.cell(ri_t, 1).value = "TOTAL"
    ws5.cell(ri_t, 2).value = total_skus_all; ws5.cell(ri_t, 2).number_format = '#,##0'
    ws5.cell(ri_t, 3).value = sum(d["muerto"] for d in prov_data.values()); ws5.cell(ri_t, 3).number_format = '#,##0'
    ws5.cell(ri_t, 4).value = sum(d["lento"]  for d in prov_data.values()); ws5.cell(ri_t, 4).number_format = '#,##0'
    ws5.cell(ri_t, 5).value = sum(d["sobre"]  for d in prov_data.values()); ws5.cell(ri_t, 5).number_format = '#,##0'
    ws5.cell(ri_t, 6).value = total_capital;  ws5.cell(ri_t, 6).number_format = '#,##0.00'
    ws5.row_dimensions[ri_t].height = 16

    wb.save(path)
    print(f"\n  ✅ Sobrestock guardado: {path}")
    print(f"     {len(rows)} SKUs totales")
    print(f"     {len(dead_rows)} stock muerto (sin ventas {DAYS}d)")
    print(f"     {len(slow_rows)} stock lento (<{SLOW_VEL_WEEKLY} u/sem, >{SLOW_DIAS_QUIEBRE}d quiebre)")
    print(f"     {len(over_rows)} sobrestock (>{OVER_DIAS_QUIEBRE}d quiebre)")
    print(f"     Capital total inmovilizado: ${total_capital:,.2f}")

# ─── ENTRYPOINT ───────────────────────────────────────────────────────────────

def main():
    print(f"\n{'═'*58}")
    print(f"  REPORTE VENTAS {DAYS} DÍAS  |  Cobertura target: {COVERAGE_DAYS}d")
    print(f"  Período: {DATE_FROM_STR} → {DATE_TO_STR}")
    print(f"  Config:  {config_path}")
    print(f"{'═'*58}\n")

    threads = []
    for key, cfg in channels.items():
        if not cfg.get("enabled", True):
            continue
        fn = worker_tn if cfg["type"] == "tiendanube" else worker_ml
        t  = threading.Thread(target=fn, args=(key, cfg), daemon=True)
        threads.append(t)
        t.start()

    for t in threads:
        t.join()

    if fetch_errors:
        print(f"\n  ⚠ Errores en canales: {fetch_errors}\n")

    print(f"\n  Construyendo tabla de análisis...")
    rows = build_rows(results)
    print(f"  {len(rows)} variantes procesadas")

    base = str(config_path.parent)
    path_repo   = args.output or f"{base}/reporte_reposicion_{DATE_TO_STR}.xlsx"
    path_sobre  = path_repo.replace("reposicion", "sobrestock") if "reposicion" in path_repo \
                  else f"{base}/reporte_sobrestock_{DATE_TO_STR}.xlsx"

    print(f"\n  Generando Excel de Reposición...")
    write_excel_reposicion(rows, path_repo)

    print(f"\n  Generando Excel de Sobrestock...")
    write_excel_sobrestock(rows, path_sobre)

    if fetch_errors:
        print(f"\n  ⚠ Canales con error (datos incompletos): {list(fetch_errors.keys())}")
    print(f"\n{'═'*58}\n")

if __name__ == "__main__":
    main()
