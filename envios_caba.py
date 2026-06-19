"""
Análisis de envíos a Capital Federal — ML (sin FULL) + Tienda Nube
Genera un Excel con resumen mensual + detalle por canal.

Uso:
    pip install requests openpyxl
    python envios_caba.py              # últimos 6 meses
    python envios_caba.py --months 12  # últimos 12 meses
"""

import requests, time, argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CREDENCIALES ──────────────────────────────────────────────────────────────
CHANNELS = {
    "tn_pret": {
        "type": "tiendanube", "label": "TN Pret a Home",
        "store_id": "2625285",
        "access_token": "7bf4cde46764d96772079d8cb1d10cd644aa35a0",
    },
    "tn_lavan": {
        "type": "tiendanube", "label": "TN Casa Lavan",
        "store_id": "6709240",
        "access_token": "e01b9ebab49d1e00973700eb3e6a1985b016e4e7",
    },
    "ml_pret": {
        "type": "mercadolibre", "label": "ML Pret a Home (sin FULL)",
        "user_id": "1255615205",
        "access_token": "APP_USR-1637574709714032-050818-ad208ec2dcc16a2434827bb0d54b524d-1255615205",
        "refresh_token": "TG-69fe5ce3cbeed400017ab4fa-1255615205",
        "client_id": "1637574709714032",
        "client_secret": "7YEH6ppegi1GbKGu6DhYJptNDeWBjq1s",
    },
    "ml_lavan": {
        "type": "mercadolibre", "label": "ML Casa Lavan (sin FULL)",
        "user_id": "189036603",
        "access_token": "APP_USR-1637574709714032-050818-c0d48ee817375be9e9c85873ed8d7cca-189036603",
        "refresh_token": "TG-69fe666782a6e000017b40b0-189036603",
        "client_id": "1637574709714032",
        "client_secret": "7YEH6ppegi1GbKGu6DhYJptNDeWBjq1s",
    },
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def is_caba(addr):
    if not addr:
        return False
    state = str(addr.get("state", addr.get("state_id", ""))).lower().strip()
    city  = str(addr.get("city", "")).lower().strip()
    zc    = str(addr.get("zip_code", addr.get("zipcode", ""))).lower().strip()
    if state in ("ar-c", "c"):
        return True
    kw = ("capital federal", "ciudad autónoma", "ciudad autonoma", "caba", "ciudad de buenos aires")
    if any(k in state for k in kw): return True
    if any(k in city  for k in kw): return True
    if zc and zc[0] == "c" and len(zc) >= 5: return True
    return False

def month_key(date_str):
    d = datetime.fromisoformat(date_str[:19])
    return f"{d.year}-{d.month:02d}"

def month_label(mk):
    y, m = mk.split("-")
    names = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    return f"{names[int(m)-1]} {y}"

# ── ML ────────────────────────────────────────────────────────────────────────
def ml_refresh(ch):
    r = requests.post("https://api.mercadolibre.com/oauth/token", data={
        "grant_type":    "refresh_token",
        "client_id":     ch["client_id"],
        "client_secret": ch["client_secret"],
        "refresh_token": ch["refresh_token"],
    })
    if r.status_code == 200:
        data = r.json()
        ch["access_token"]  = data["access_token"]
        ch["refresh_token"] = data["refresh_token"]
        print("  Token renovado OK")
    else:
        print(f"  WARN: no se pudo renovar token: {r.status_code}")

def fetch_ml_orders(ch, date_from):
    ml_refresh(ch)
    token, uid = ch["access_token"], ch["user_id"]
    orders, offset = [], 0
    while True:
        url = (f"https://api.mercadolibre.com/orders/search"
               f"?seller={uid}&order.status=paid&sort=date_desc"
               f"&offset={offset}&limit=50"
               f"&order.date_created.from={date_from}")
        r = requests.get(url, headers={"Authorization": f"Bearer {token}"})
        if r.status_code == 429: time.sleep(2); continue
        if r.status_code != 200:
            print(f"  ERROR ML {r.status_code}: {r.text[:120]}"); break
        data    = r.json()
        results = data.get("results", [])
        orders.extend(results)
        total = data.get("paging", {}).get("total", 0)
        print(f"  Descargando {len(orders)}/{total}...", end="\r")
        offset += 50
        if offset >= min(total, 1000): break
        time.sleep(0.35)
    print(f"  {len(orders)} ordenes ML descargadas        ")
    return orders

def process_ml(orders):
    result = []
    for o in orders:
        shipping = o.get("shipping") or {}
        tags     = o.get("tags") or []
        logtype  = shipping.get("logistic_type", "")
        if logtype == "fulfillment" or "fulfillment" in tags:
            continue
        addr = (shipping.get("receiver_address") or
                (o.get("buyer") or {}).get("shipping_address") or {})
        if not is_caba(addr):
            continue
        total        = float(o.get("total_amount") or 0)
        total_w_ship = float(o.get("total_amount_with_shipping") or total)
        ship_cost    = round(total_w_ship - total, 2) if total_w_ship > total else 0
        result.append({
            "id":       str(o["id"]),
            "date":     o["date_created"][:10],
            "month":    month_key(o["date_created"]),
            "total":    total,
            "ship_cost":ship_cost,
            "logistic": logtype or "normal",
            "city":     addr.get("city", ""),
            "state":    addr.get("state_id", ""),
        })
    return result

# ── TN ────────────────────────────────────────────────────────────────────────
def fetch_tn_orders(ch, date_from):
    sid, token = ch["store_id"], ch["access_token"]
    headers = {
        "Authentication": f"bearer {token}",
        "User-Agent": "PretaHome Analytics (admin@pretahome.com.ar)",
    }
    orders, page = [], 1
    while True:
        url = (f"https://api.tiendanube.com/v1/{sid}/orders"
               f"?page={page}&per_page=200"
               f"&created_at_min={date_from}"
               f"&payment_status=paid,authorized"
               f"&fields=id,created_at,status,total,shipping_address,"
               f"shipping_cost_owner,shipping_cost_customer")
        r = requests.get(url, headers=headers)
        if r.status_code == 401:
            print("  ERROR TN 401: token expirado"); break
        if r.status_code == 429:
            time.sleep(1); continue
        if r.status_code != 200:
            print(f"  ERROR TN {r.status_code}: {r.text[:120]}"); break
        data = r.json()
        if not isinstance(data, list) or not data:
            break
        valid = [o for o in data if o.get("status") != "cancelled"]
        orders.extend(valid)
        print(f"  Descargando {len(orders)}...", end="\r")
        if len(data) < 200: break
        page += 1
        time.sleep(0.6)
    print(f"  {len(orders)} ordenes TN descargadas        ")
    return orders

def process_tn(orders):
    result = []
    for o in orders:
        addr = o.get("shipping_address") or {}
        if not is_caba(addr):
            continue
        result.append({
            "id":       str(o["id"]),
            "date":     o["created_at"][:10],
            "month":    month_key(o["created_at"]),
            "total":    float(o.get("total") or 0),
            "ship_cost":float(o.get("shipping_cost_owner") or 0),
            "logistic": "tiendanube",
            "city":     addr.get("city", ""),
            "state":    addr.get("province", ""),
        })
    return result

# ── EXCEL ─────────────────────────────────────────────────────────────────────
def make_excel(all_data, totals_fetched, months_back, output_path):
    wb = Workbook()
    thin = Side(style="thin", color="BDC9D9")
    BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)

    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    ALT_FILL  = PatternFill("solid", fgColor="EEF3F9")
    TOT_FILL  = PatternFill("solid", fgColor="D9E2F3")
    GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")
    SUB_FILL  = PatternFill("solid", fgColor="2E75B6")
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BLUE_FONT = Font(name="Arial", size=10, color="0000FF")
    BOLD      = Font(name="Arial", bold=True, size=10)
    NORM      = Font(name="Arial", size=10)
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RGT  = Alignment(horizontal="right",  vertical="center")
    LFT  = Alignment(horizontal="left",   vertical="center")

    def h(ws, row, col, val, fill=HDR_FILL):
        c = ws.cell(row=row, column=col, value=val)
        c.fill=fill; c.font=HDR_FONT; c.alignment=CTR; c.border=BRD

    def cd(ws, row, col, val, bold=False, align=RGT, fill=None, nf=None, font=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = font or Font(name="Arial", bold=bold, size=10)
        c.alignment = align; c.border = BRD
        if fill: c.fill = fill
        if nf:   c.number_format = nf
        return c

    # ── Hoja 1: Resumen mensual ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen mensual"
    ws.freeze_panes = "A3"

    all_orders = []
    for k, orders in all_data.items():
        for o in orders:
            o["canal"] = k
            all_orders.append(o)

    months  = sorted(set(o["month"] for o in all_orders))
    canales = list(all_data.keys())
    nm      = len(months) or 1

    # Título
    ncols = 3 + len(canales) * 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=f"Envíos a Capital Federal — Últimos {months_back} meses  |  Generado {datetime.now().strftime('%d/%m/%Y')}")
    t.font=Font(name="Arial", bold=True, size=12, color="FFFFFF"); t.fill=HDR_FILL; t.alignment=CTR

    col = 1
    h(ws, 2, col, "Mes"); col += 1
    for k in canales:
        h(ws, 2, col, CHANNELS[k]["label"] + "\nEnvíos"); col += 1
        h(ws, 2, col, CHANNELS[k]["label"] + "\nCosto envío"); col += 1
    h(ws, 2, col, "TOTAL\nEnvíos"); col += 1
    h(ws, 2, col, "TOTAL\nCosto"); col += 1
    h(ws, 2, col, "Costo prom.\npor envío")
    ws.row_dimensions[2].height = 32

    gt_envios = gt_costo = 0
    for i, mk in enumerate(months):
        fill = ALT_FILL if i % 2 == 0 else None
        col  = 1
        cd(ws, 3+i, col, month_label(mk), bold=True, align=LFT, fill=fill); col += 1
        te = tc = 0
        for k in canales:
            env = sum(1 for o in all_orders if o["canal"]==k and o["month"]==mk)
            cst = sum(o["ship_cost"] for o in all_orders if o["canal"]==k and o["month"]==mk)
            cd(ws, 3+i, col, env, fill=fill, align=CTR); col += 1
            cd(ws, 3+i, col, cst, fill=fill, nf='"$"#,##0'); col += 1
            te += env; tc += cst
        cd(ws, 3+i, col, te, bold=True, fill=fill, align=CTR); col += 1
        cd(ws, 3+i, col, tc, bold=True, fill=fill, nf='"$"#,##0'); col += 1
        cd(ws, 3+i, col, round(tc/te,0) if te else 0, fill=fill, nf='"$"#,##0')
        gt_envios += te; gt_costo += tc

    # Total
    tr = 3 + len(months)
    col = 1
    cd(ws, tr, col, "TOTAL", bold=True, align=LFT, fill=TOT_FILL); col += 1
    for k in canales:
        env = sum(1 for o in all_orders if o["canal"]==k)
        cst = sum(o["ship_cost"] for o in all_orders if o["canal"]==k)
        cd(ws, tr, col, env, bold=True, fill=TOT_FILL, align=CTR); col += 1
        cd(ws, tr, col, cst, bold=True, fill=TOT_FILL, nf='"$"#,##0'); col += 1
    cd(ws, tr, col, gt_envios, bold=True, fill=TOT_FILL, align=CTR); col += 1
    cd(ws, tr, col, gt_costo,  bold=True, fill=TOT_FILL, nf='"$"#,##0'); col += 1
    cd(ws, tr, col, round(gt_costo/gt_envios,0) if gt_envios else 0, bold=True, fill=TOT_FILL, nf='"$"#,##0')

    # Promedio mensual
    pr = tr + 1
    col = 1
    cd(ws, pr, col, "Promedio / mes", bold=True, align=LFT); col += 1
    for k in canales:
        env = sum(1 for o in all_orders if o["canal"]==k)
        cst = sum(o["ship_cost"] for o in all_orders if o["canal"]==k)
        cd(ws, pr, col, round(env/nm,1), align=CTR); col += 1
        cd(ws, pr, col, round(cst/nm,0), nf='"$"#,##0'); col += 1
    cd(ws, pr, col, round(gt_envios/nm,1), bold=True, align=CTR); col += 1
    cd(ws, pr, col, round(gt_costo/nm,0),  bold=True, nf='"$"#,##0'); col += 1
    cd(ws, pr, col, round(gt_costo/gt_envios,0) if gt_envios else 0, bold=True, nf='"$"#,##0')

    ws.column_dimensions["A"].width = 14
    for i in range(2, ncols+1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # ── Hoja 2: Detalle órdenes ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle órdenes")
    ws2.freeze_panes = "A2"
    hdrs2 = ["Canal","ID Orden","Fecha","Mes","Ciudad","Prov/Estado","Logística","Total venta","Costo envío"]
    for ci, hv in enumerate(hdrs2, 1):
        h(ws2, 1, ci, hv)
    for i, o in enumerate(sorted(all_orders, key=lambda x: x["date"], reverse=True)):
        fill = ALT_FILL if i%2==0 else None
        r2   = i + 2
        cd(ws2, r2, 1, CHANNELS[o["canal"]]["label"], align=LFT, fill=fill)
        cd(ws2, r2, 2, o["id"],               align=LFT,  fill=fill)
        cd(ws2, r2, 3, o["date"],             align=CTR,  fill=fill)
        cd(ws2, r2, 4, month_label(o["month"]),align=CTR, fill=fill)
        cd(ws2, r2, 5, o.get("city",""),      align=LFT,  fill=fill)
        cd(ws2, r2, 6, o.get("state",""),     align=LFT,  fill=fill)
        cd(ws2, r2, 7, o.get("logistic",""),  align=LFT,  fill=fill)
        cd(ws2, r2, 8, o["total"],            fill=fill,  nf='"$"#,##0')
        cd(ws2, r2, 9, o["ship_cost"],        fill=fill,  nf='"$"#,##0')
    for ci, w in enumerate([22,14,12,12,18,20,14,14,14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Hoja 3: Análisis camioneta ──────────────────────────────────────────
    ws3 = wb.create_sheet("Análisis camioneta")
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 32

    def h3(r, c, v, fill=SUB_FILL):
        cx = ws3.cell(row=r, column=c, value=v)
        cx.fill=fill; cx.font=HDR_FONT; cx.alignment=LFT; cx.border=BRD

    def d3(r, c, v, bold=False, nf=None, fill=None, font=None, align=RGT):
        cx = ws3.cell(row=r, column=c, value=v)
        cx.font = font or Font(name="Arial", bold=bold, size=10)
        cx.alignment=align; cx.border=BRD
        if nf:   cx.number_format=nf
        if fill: cx.fill=fill
        return cx

    ws3.merge_cells("A1:C1")
    t3 = ws3.cell(row=1, column=1, value="¿Conviene comprar una camioneta para envíos CABA?")
    t3.font=Font(name="Arial", bold=True, size=13, color="FFFFFF"); t3.fill=HDR_FILL; t3.alignment=CTR

    r = 3
    h3(r,1,"Datos actuales (de APIs)"); h3(r,2,"Valor"); h3(r,3,"Notas"); r+=1

    prom_env = round(gt_envios/nm, 1)
    prom_cst = round(gt_costo/nm, 0)
    prom_xenv= round(gt_costo/gt_envios, 0) if gt_envios else 0

    datos = [
        ("Meses analizados",                 nm,          ""),
        ("Total envíos CABA (período)",      gt_envios,   "Sin FULL de ML"),
        ("Promedio envíos CABA / mes",       prom_env,    ""),
        ("Total costo envíos (período)",     gt_costo,    "Lo que pagaste en envíos"),
        ("Promedio costo envíos / mes",      prom_cst,    ""),
        ("Costo promedio por envío",         prom_xenv,   ""),
    ]
    for label, val, note in datos:
        nf = '"$"#,##0' if "costo" in label.lower() else None
        d3(r,1,label,align=LFT); d3(r,2,val,nf=nf)
        ws3.cell(row=r,column=3,value=note).font=Font(name="Arial",size=9,color="666666")
        ws3.cell(row=r,column=3).alignment=LFT; ws3.cell(row=r,column=3).border=BRD
        r+=1

    r+=1
    h3(r,1,"Supuestos (editá los valores azules)"); h3(r,2,"Valor"); h3(r,3,""); r+=1
    sup_start = r

    supuestos = [
        ("Precio camioneta (ARS)",            3_500_000, '"$"#,##0'),
        ("Años de amortización",              5,         "0"),
        ("Seguro mensual (ARS)",              120_000,   '"$"#,##0'),
        ("Costo combustible por entrega (ARS)",2_500,    '"$"#,##0'),
        ("Entregas por día de reparto",       8,         "0"),
        ("Días de reparto por mes",           4,         "0"),
        ("Sueldo repartidor por hora (ARS)",  3_000,     '"$"#,##0'),
        ("Horas por día de reparto",          6,         "0"),
        ("Mantenimiento mensual (ARS)",       50_000,    '"$"#,##0'),
    ]
    for label, val, nf in supuestos:
        d3(r,1,label,align=LFT)
        d3(r,2,val,nf=nf,font=BLUE_FONT)
        ws3.cell(row=r,column=3).border=BRD
        r+=1

    r+=1
    h3(r,1,"Costo mensual camioneta (calculado)"); h3(r,2,"Valor"); h3(r,3,"Fórmula"); r+=1
    calc_start = r
    s = sup_start  # alias corto

    calcs = [
        ("Amortización mensual",     f"=B{s}/(B{s+1}*12)",         "precio / (años × 12)"),
        ("Combustible mensual",      f"=B{s+3}*B{s+4}*B{s+5}",     "costo/entrega × entregas/día × días/mes"),
        ("Costo personal mensual",   f"=B{s+6}*B{s+7}*B{s+5}",     "sueldo/hora × horas × días/mes"),
        ("Seguro mensual",           f"=B{s+2}",                    ""),
        ("Mantenimiento mensual",    f"=B{s+8}",                    ""),
    ]
    for label, formula, note in calcs:
        d3(r,1,label,align=LFT)
        d3(r,2,formula,nf='"$"#,##0')
        ws3.cell(row=r,column=3,value=note).font=Font(name="Arial",size=9,color="666666")
        ws3.cell(row=r,column=3).alignment=LFT; ws3.cell(row=r,column=3).border=BRD
        r+=1

    total_cam_row = r
    d3(r,1,"COSTO TOTAL MENSUAL CAMIONETA",bold=True,align=LFT,fill=TOT_FILL)
    d3(r,2,f"=SUM(B{calc_start}:B{r-1})",bold=True,nf='"$"#,##0',fill=TOT_FILL)
    ws3.cell(row=r,column=3).border=BRD; ws3.cell(row=r,column=3).fill=TOT_FILL
    r+=2

    h3(r,1,"Comparativa"); h3(r,2,"Valor"); h3(r,3,""); r+=1
    d3(r,1,"Gasto actual envíos/mes (histórico)",align=LFT)
    d3(r,2,prom_cst,nf='"$"#,##0'); ws3.cell(row=r,column=3).border=BRD; r+=1
    d3(r,1,"Costo total camioneta/mes",align=LFT)
    d3(r,2,f"=B{total_cam_row}",nf='"$"#,##0'); ws3.cell(row=r,column=3).border=BRD; r+=1

    d3(r,1,"Ahorro mensual  (+ahorro / −sobrecosto)",bold=True,align=LFT)
    d3(r,2,f"=B{r-2}-B{r-1}",bold=True,nf='"$"#,##0',fill=GRN_FILL)
    ws3.cell(row=r,column=3).border=BRD; r+=1

    d3(r,1,"Costo por envío con camioneta",align=LFT)
    formula_cpu = f"=IFERROR(B{total_cam_row}/{prom_env},\"—\")" if prom_env else "—"
    d3(r,2,formula_cpu,nf='"$"#,##0'); ws3.cell(row=r,column=3).border=BRD; r+=1

    d3(r,1,"Costo por envío actual (histórico)",align=LFT)
    d3(r,2,prom_xenv,nf='"$"#,##0'); ws3.cell(row=r,column=3).border=BRD; r+=1

    d3(r,1,"Envíos/mes mínimos para que la camioneta se pague sola",bold=True,align=LFT)
    min_formula = f"=IFERROR(ROUNDUP(B{total_cam_row}/{prom_xenv},0),\"—\")" if prom_xenv else "—"
    d3(r,2,min_formula,bold=True,align=CTR,fill=TOT_FILL)
    ws3.cell(row=r,column=3,value="Si enviás más que esto por mes, conviene la camioneta").font=Font(name="Arial",size=9,color="666666")
    ws3.cell(row=r,column=3).alignment=LFT; ws3.cell(row=r,column=3).border=BRD

    wb.save(output_path)
    print(f"\nExcel guardado en: {output_path}")

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--months", type=int, default=6)
    args = parser.parse_args()

    date_from = (datetime.now() - timedelta(days=args.months*30)).strftime("%Y-%m-%dT00:00:00-03:00")
    print(f"Periodo: ultimos {args.months} meses (desde {date_from[:10]})\n")

    all_data = {}
    totals   = {}

    for key, ch in CHANNELS.items():
        print(f"{'─'*50}\n{ch['label']}")
        if ch["type"] == "tiendanube":
            raw = fetch_tn_orders(ch, date_from); orders = process_tn(raw)
        else:
            raw = fetch_ml_orders(ch, date_from); orders = process_ml(raw)
        print(f"  → {len(orders)} envios a CABA (de {len(raw)} totales)")
        all_data[key] = orders
        totals[key]   = len(raw)

    total_caba = sum(len(v) for v in all_data.values())
    print(f"\n{'─'*50}\nTOTAL envios CABA: {total_caba}")
    for k,v in all_data.items():
        pct = f"{len(v)/totals[k]*100:.1f}%" if totals.get(k) else "—"
        print(f"  {CHANNELS[k]['label']}: {len(v)} ({pct} del total)")

    output = f"envios_caba_{datetime.now().strftime('%Y%m%d')}.xlsx"
    make_excel(all_data, totals, args.months, output)
