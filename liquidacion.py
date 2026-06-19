#!/usr/bin/env python3
"""
liquidacion.py — Reporte de liquidación: productos en stock sin ventas

Genera un Excel con:
- Todos los productos activos en TN con stock > 0
- Días sin ventas (desde última venta o desde publicación)
- Costo, precio, stock, margen
- Links a ML y TN
- Ordenado por días sin ventas (mayor a menor)

Uso:
    python3 liquidacion.py
    python3 liquidacion.py --output mi_reporte.xlsx
    python3 liquidacion.py --min-dias 30   # solo productos sin ventas hace +30d
"""

import json, argparse
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

ROOT     = Path(__file__).parent
DATA_DIR = ROOT / "data"

# ─── CARGA ────────────────────────────────────────────────────────────────────

def load_cache(canal):
    p = DATA_DIR / f"cache_{canal}.json"
    if not p.exists():
        return None
    with open(p) as f:
        return json.load(f)

def load_config():
    with open(ROOT / "config.json") as f:
        return json.load(f)

# ─── LÓGICA ───────────────────────────────────────────────────────────────────

def days_since(date_str):
    if not date_str or date_str == "—":
        return None
    try:
        d = datetime.strptime(date_str[:10], "%Y-%m-%d").date()
        return (date.today() - d).days
    except:
        return None

def build_report(min_dias=0):
    config = load_config()

    # ── Catálogo TN Pret (fuente de verdad: stock, costo, precio) ──────────────
    tn_pret = load_cache("tn_pret")
    if not tn_pret:
        print("ERROR: No hay cache de tn_pret. Corré sync.py primero.")
        exit(1)
    tn_items = tn_pret.get("items", {})

    # ── Últimas ventas por SKU (de todos los canales) ──────────────────────────
    last_sale_by_sku = {}   # sku -> fecha última venta

    for canal_key in ["ml_pret", "tn_pret", "tn_lavan"]:
        cache = load_cache(canal_key)
        if not cache:
            continue
        items  = cache.get("items", {})
        orders = cache.get("orders", [])
        is_ml  = canal_key.startswith("ml")

        for order in orders:
            if not isinstance(order, dict):
                continue
            date_str = order.get("date_created" if is_ml else "created_at", "")[:10]
            if not date_str:
                continue
            lines = order.get("order_items" if is_ml else "products", [])
            for line in lines:
                if not isinstance(line, dict):
                    continue
                if is_ml:
                    iid    = line.get("item", {}).get("id", "")
                    var_id = str(line.get("item", {}).get("variation_id") or "")
                    item   = items.get(iid, {})
                    var_skus = item.get("var_skus", {})
                    if var_id and var_skus.get(var_id):
                        sku = var_skus[var_id]
                    elif item.get("sku"):
                        sku = item["sku"]
                    else:
                        sku = iid
                else:
                    vid  = str(line.get("variant_id") or line.get("product_id") or "")
                    item = items.get(vid, {})
                    sku  = item.get("sku") or vid

                if not sku:
                    continue
                prev = last_sale_by_sku.get(sku)
                if prev is None or date_str > prev:
                    last_sale_by_sku[sku] = date_str

    # ── Links ML por SKU ───────────────────────────────────────────────────────
    ml_links_by_sku = defaultdict(list)   # sku -> [permalink, ...]
    ml_cache = load_cache("ml_pret")
    if ml_cache:
        ml_items = ml_cache.get("items", {})
        for iid, item in ml_items.items():
            sku = item.get("sku") or ""
            var_skus = item.get("var_skus", {})
            skus_in_item = set()
            if sku:
                skus_in_item.add(sku)
            skus_in_item.update(var_skus.values())
            permalink = item.get("permalink", "")
            if permalink:
                for s in skus_in_item:
                    if s not in ml_links_by_sku or permalink not in ml_links_by_sku[s]:
                        ml_links_by_sku[s].append(permalink)

    # ── Construir filas ────────────────────────────────────────────────────────
    rows = []
    today = date.today()

    for vid, item in tn_items.items():
        stock = item.get("stock", 0) or 0
        if stock <= 0:
            continue  # Solo con stock

        sku        = item.get("sku") or vid
        title      = item.get("title", "")
        price      = float(item.get("price", 0) or 0)
        cost       = float(item.get("cost", 0) or 0)
        created_at = item.get("created_at", "") or ""
        margen     = round((price - cost) / price * 100, 1) if price > 0 and cost > 0 else None

        last_sale  = last_sale_by_sku.get(sku)
        ds_venta   = days_since(last_sale)
        ds_pub     = days_since(created_at) if created_at else None

        # Días sin ventas: desde la última venta, o desde publicación si nunca vendió
        if ds_venta is not None:
            dias_sin_ventas = ds_venta
            razon           = f"Última venta: {last_sale}"
        elif ds_pub is not None:
            dias_sin_ventas = ds_pub
            razon           = f"Publicado: {created_at[:10]} (nunca vendió)"
        else:
            dias_sin_ventas = 9999
            razon           = "Sin datos de venta ni fecha de publicación"

        if min_dias > 0 and dias_sin_ventas < min_dias:
            continue

        # Clasificación
        if dias_sin_ventas >= 90:
            clasificacion = "🔴 +90 días"
        elif dias_sin_ventas >= 60:
            clasificacion = "🟠 +60 días"
        elif dias_sin_ventas >= 30:
            clasificacion = "🟡 +30 días"
        else:
            clasificacion = "🟢 <30 días"

        # Links ML
        ml_links = ml_links_by_sku.get(sku, [])
        ml_link1 = ml_links[0] if len(ml_links) > 0 else ""
        ml_link2 = ml_links[1] if len(ml_links) > 1 else ""

        # Link TN (buscar en tn_pret por product_id)
        tn_store_id = config.get("channels", {}).get("tn_pret", {}).get("store_id", "")
        product_id  = item.get("product_id", "")
        tn_link     = f"https://www.{tn_store_id}.mitiendanube.com/productos/{product_id}" if product_id else ""

        rows.append({
            "SKU":               sku,
            "Producto":          title,
            "Días sin ventas":   dias_sin_ventas if dias_sin_ventas < 9999 else None,
            "Clasificación":     clasificacion,
            "Última venta":      last_sale or "Nunca",
            "Fecha publicación": created_at[:10] if created_at else "",
            "Stock actual":      stock,
            "Precio ($)":        round(price),
            "Costo ($)":         round(cost) if cost > 0 else "",
            "Margen (%)":        margen if margen is not None else "",
            "Revenue potencial": round(price * stock),
            "Costo total stock": round(cost * stock) if cost > 0 else "",
            "En ML":             "Sí" if ml_links else "No",
            "Link ML 1":         ml_link1,
            "Link ML 2":         ml_link2,
            "Link TN":           tn_link,
            "Razón":             razon,
        })

    # Ordenar por días sin ventas (mayor a menor)
    rows.sort(key=lambda r: r["Días sin ventas"] or 0, reverse=True)
    return rows

# ─── EXCEL ────────────────────────────────────────────────────────────────────

def generate_excel(rows, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.styles.differential import DifferentialStyle

    wb = Workbook()

    # ── HOJA 1: Todos los productos ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Liquidación"

    # Colores
    C_HEADER_BG  = "1F2937"
    C_HEADER_FG  = "F9FAFB"
    C_RED_BG     = "FEE2E2"
    C_ORANGE_BG  = "FFEDD5"
    C_YELLOW_BG  = "FEF9C3"
    C_GREEN_BG   = "DCFCE7"
    C_ZEBRA      = "F8FAFC"
    C_BORDER     = "E5E7EB"
    C_ACCENT     = "DC2626"

    thin = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = list(rows[0].keys()) if rows else []
    col_widths = {
        "SKU": 22, "Producto": 48, "Días sin ventas": 16, "Clasificación": 14,
        "Última venta": 14, "Fecha publicación": 16, "Stock actual": 12,
        "Precio ($)": 13, "Costo ($)": 12, "Margen (%)": 11,
        "Revenue potencial": 18, "Costo total stock": 17,
        "En ML": 8, "Link ML 1": 55, "Link ML 2": 55, "Link TN": 45, "Razón": 40,
    }

    # Título
    ws.merge_cells("A1:Q1")
    title_cell = ws["A1"]
    title_cell.value = f"Reporte de Liquidación — Pret a Home / Casa Lavan — {date.today().strftime('%d/%m/%Y')}"
    title_cell.font      = Font(name="Arial", bold=True, size=13, color=C_HEADER_FG)
    title_cell.fill      = PatternFill("solid", fgColor=C_ACCENT)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtítulo con resumen
    r90  = sum(1 for r in rows if r["Días sin ventas"] and r["Días sin ventas"] >= 90)
    r60  = sum(1 for r in rows if r["Días sin ventas"] and 60 <= r["Días sin ventas"] < 90)
    r30  = sum(1 for r in rows if r["Días sin ventas"] and 30 <= r["Días sin ventas"] < 60)
    total_rev = sum(r["Revenue potencial"] for r in rows)
    ws.merge_cells("A2:Q2")
    sub = ws["A2"]
    sub.value = f"Total: {len(rows)} SKUs con stock  |  🔴 +90d: {r90}  |  🟠 +60d: {r60}  |  🟡 +30d: {r30}  |  Revenue potencial total: ${total_rev:,.0f}"
    sub.font      = Font(name="Arial", size=10, color="374151")
    sub.fill      = PatternFill("solid", fgColor="FEF3C7")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(header, 14)
    ws.row_dimensions[3].height = 30

    # Datos
    color_map = {
        "🔴 +90 días": C_RED_BG,
        "🟠 +60 días": C_ORANGE_BG,
        "🟡 +30 días": C_YELLOW_BG,
        "🟢 <30 días": C_GREEN_BG,
    }

    for row_idx, row in enumerate(rows, 4):
        clasi      = row.get("Clasificación", "")
        bg_color   = color_map.get(clasi, C_ZEBRA if row_idx % 2 == 0 else "FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            val  = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = PatternFill("solid", fgColor=bg_color)
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Números alineados a la derecha
            if header in ("Días sin ventas", "Stock actual", "Precio ($)", "Costo ($)",
                          "Margen (%)", "Revenue potencial", "Costo total stock"):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if header == "Margen (%)" and val != "":
                    cell.number_format = "0.0%"
                elif header in ("Precio ($)", "Costo ($)", "Revenue potencial", "Costo total stock"):
                    cell.number_format = '#,##0'

            # Links clickeables
            if header in ("Link ML 1", "Link ML 2", "Link TN") and val:
                cell.hyperlink = val
                cell.value     = "Ver publicación →"
                cell.font      = Font(name="Arial", size=9, color="1D4ED8", underline="single")

            # Días sin ventas — negrita si es alto
            if header == "Días sin ventas" and val and val >= 90:
                cell.font = Font(name="Arial", size=9, bold=True, color="991B1B")

        ws.row_dimensions[row_idx].height = 16

    # Freeze panes
    ws.freeze_panes = "A4"

    # Autofilter
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"

    # ── HOJA 2: Resumen por umbral ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 20

    ws2["A1"] = "Resumen de liquidación"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color=C_HEADER_FG)
    ws2["A1"].fill = PatternFill("solid", fgColor=C_ACCENT)
    ws2.merge_cells("A1:D1")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 26

    for col, label in enumerate(["Categoría", "Cant. SKUs", "Revenue potencial", "Costo total stock"], 1):
        c = ws2.cell(row=2, column=col, value=label)
        c.font = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
        c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center")

    groups = [
        ("🔴 +90 días (urgente)",  [r for r in rows if r["Días sin ventas"] and r["Días sin ventas"] >= 90]),
        ("🟠 +60 días",            [r for r in rows if r["Días sin ventas"] and 60 <= r["Días sin ventas"] < 90]),
        ("🟡 +30 días",            [r for r in rows if r["Días sin ventas"] and 30 <= r["Días sin ventas"] < 60]),
        ("🟢 <30 días",            [r for r in rows if not r["Días sin ventas"] or r["Días sin ventas"] < 30]),
        ("TOTAL",                  rows),
    ]

    bg_map = {
        "🔴 +90 días (urgente)": C_RED_BG,
        "🟠 +60 días":            C_ORANGE_BG,
        "🟡 +30 días":            C_YELLOW_BG,
        "🟢 <30 días":            C_GREEN_BG,
        "TOTAL":                  "E5E7EB",
    }

    for i, (label, group) in enumerate(groups, 3):
        rev  = sum(r["Revenue potencial"] for r in group)
        cost = sum(r["Costo total stock"] for r in group if isinstance(r["Costo total stock"], (int, float)))
        bg   = bg_map.get(label, "FFFFFF")
        bold = label == "TOTAL"
        for col, val in enumerate([label, len(group), rev, cost], 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.font = Font(name="Arial", bold=bold, size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = border
            c.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
            if col in (3, 4):
                c.number_format = "$#,##0"
        ws2.row_dimensions[i].height = 20

    wb.save(output_path)
    print(f"  Guardado: {output_path} ({Path(output_path).stat().st_size//1024} KB)")

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output",   default="liquidacion.xlsx")
    parser.add_argument("--min-dias", type=int, default=0,
                        help="Solo incluir productos con X+ días sin ventas")
    args = parser.parse_args()

    print(f"\n  Reporte de liquidación — {date.today().strftime('%d/%m/%Y')}")
    print(f"  {'─'*48}")

    print("  Procesando datos...")
    rows = build_report(min_dias=args.min_dias)
    print(f"  {len(rows)} SKUs con stock encontrados")

    r90 = sum(1 for r in rows if r["Días sin ventas"] and r["Días sin ventas"] >= 90)
    r60 = sum(1 for r in rows if r["Días sin ventas"] and 60 <= (r["Días sin ventas"] or 0) < 90)
    r30 = sum(1 for r in rows if r["Días sin ventas"] and 30 <= (r["Días sin ventas"] or 0) < 60)
    print(f"  🔴 +90 días: {r90}  |  🟠 +60 días: {r60}  |  🟡 +30 días: {r30}")

    print("  Generando Excel...")
    generate_excel(rows, args.output)

    print(f"\n  Abrilo con: open {args.output}\n")

if __name__ == "__main__":
    main()
