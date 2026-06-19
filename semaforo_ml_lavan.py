#!/usr/bin/env python3
"""
semaforo_ml_lavan.py — Semáforo de publicaciones ML Casa Lavan
Uso: python3 semaforo_ml.py
"""

import json, requests, time, argparse, sys
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from pathlib import Path

parser = argparse.ArgumentParser()
parser.add_argument("--days",   type=int, default=10)
parser.add_argument("--config", type=str, default=None)
args = parser.parse_args()

DAYS      = 7
NOW       = datetime.now(timezone.utc)

# Calcular el último lunes-domingo completo
# Hoy es weekday(): 0=lunes ... 6=domingo
today = NOW.date()
# Último domingo
last_sunday  = today - timedelta(days=(today.weekday() + 1) % 7)
# Lunes de esa semana
last_monday  = last_sunday - timedelta(days=6)

DATE_FROM = datetime(last_monday.year, last_monday.month, last_monday.day,
                     0, 0, 0, tzinfo=timezone.utc)
DATE_TO   = datetime(last_sunday.year, last_sunday.month, last_sunday.day,
                     23, 59, 59, tzinfo=timezone.utc)

DATE_FROM_SIMPLE = last_monday.strftime("%Y-%m-%d")
DATE_TO_SIMPLE   = last_sunday.strftime("%Y-%m-%d")
DATE_FROM_STR    = DATE_FROM.strftime("%Y-%m-%dT%H:%M:%S.000-00:00")
DATE_TO_STR      = DATE_TO.strftime("%Y-%m-%dT%H:%M:%S.000-00:00")

def find_config():
    if args.config:
        p = Path(args.config)
        if p.exists(): return p
        sys.exit(f"Config no encontrado: {args.config}")
    for p in [
        Path(__file__).parent / "config.json",
        Path.home() / "Desktop" / "ecommerce-bi" / "config.json",
    ]:
        if p.exists(): return p
    sys.exit("No se encontró config.json")

config_path = find_config()
with open(config_path) as f:
    CONFIG = json.load(f)
cfg_lavan = CONFIG["channels"]["ml_lavan"]

def log(msg): print(f"  {msg}", flush=True)

# ─── ML ──────────────────────────────────────────────────────────────────────

def ml_refresh(cfg):
    r = requests.post("https://api.mercadolibre.com/oauth/token", data={
        "grant_type": "refresh_token",
        "client_id": cfg["client_id"],
        "client_secret": cfg["client_secret"],
        "refresh_token": cfg["refresh_token"],
    }, timeout=15)
    if r.status_code == 200:
        return r.json()["access_token"]
    return cfg["access_token"]

def ml_get(url, token, params=None, retries=4):
    headers = {"Authorization": f"Bearer {token}"}
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=20)
            if r.status_code == 200:
                return r.json()
            if r.status_code in (429, 500):
                time.sleep(3 * (attempt + 1))
                continue
            return None
        except Exception:
            time.sleep(2)
    return None

# ─── ITEMS — scroll para traer los 3929 sin límite ───────────────────────────

def get_all_item_ids(token, user_id):
    """
    Usa search_type=scan + scroll_id para traer TODOS los items activos.
    ML permite máx offset=1000 en paginación normal, pero scroll no tiene límite.
    """
    item_ids  = []
    scroll_id = None
    total     = None
    page      = 0

    while True:
        params = {"limit": 100, "status": "active", "search_type": "scan"}
        if scroll_id:
            params["scroll_id"] = scroll_id

        data = ml_get(
            f"https://api.mercadolibre.com/users/{user_id}/items/search",
            token, params=params
        )
        if not data:
            break

        if total is None:
            total = data.get("paging", {}).get("total", 0)
            log(f"  Total publicaciones activas en ML: {total}")

        batch = data.get("results", [])
        if not batch:
            break

        item_ids.extend(batch)
        scroll_id = data.get("scroll_id")
        page += 1

        if page % 5 == 0 or len(item_ids) >= total:
            log(f"  Items traídos: {len(item_ids)}/{total}")

        if len(item_ids) >= total or not scroll_id:
            break

        time.sleep(0.3)

    log(f"  ✓ Total items recuperados: {len(item_ids)}/{total}")
    return list(dict.fromkeys(item_ids))

# ─── DETALLE DE ITEMS ─────────────────────────────────────────────────────────

def get_items_detail(token, item_ids):
    details = {}
    total   = len(item_ids)
    for i in range(0, total, 20):
        batch = item_ids[i:i+20]
        data  = ml_get("https://api.mercadolibre.com/items", token,
                        params={"ids": ",".join(batch)})
        if data:
            for entry in data:
                if entry.get("code") == 200:
                    body = entry["body"]
                    # SKU: buscar en seller_custom_field o en atributos
                    sku = body.get("seller_custom_field") or ""
                    if not sku:
                        for attr in body.get("attributes", []):
                            if attr.get("id") in ("SELLER_SKU", "SKU"):
                                sku = attr.get("value_name") or ""
                                break

                    details[body["id"]] = {
                        "title":        body.get("title", ""),
                        "sku":          sku,
                        "price":        body.get("price", 0),
                        "permalink":    body.get("permalink") or "",
                        "status":       body.get("status", ""),
                        "stock":        body.get("available_quantity", 0),
                        "logistic_type": body.get("shipping", {}).get("logistic_type", ""),
                    }
        if (i // 20 + 1) % 50 == 0:
            log(f"  Detalle: {min(i+20, total)}/{total} items")
        time.sleep(0.25)
    return details

# ─── VISITAS — batch de 50 con fecha simple ───────────────────────────────────

def get_visits(token, item_ids):
    """
    /items/{id}/visits/time_window?last=7&unit=day
    Funciona correctamente. Usamos threads paralelos para bajar de 13min a ~1min.
    """
    visits  = {}
    total   = len(item_ids)
    lock    = __import__('threading').Lock()
    counter = [0]

    import threading

    def fetch_one(item_id):
        data = ml_get(
            f"https://api.mercadolibre.com/items/{item_id}/visits/time_window",
            token, params={"last": DAYS, "unit": "day"}
        )
        with lock:
            if data and isinstance(data, dict):
                visits[item_id] = data.get("total_visits", 0)
            else:
                visits[item_id] = 0
            counter[0] += 1
            if counter[0] % 200 == 0:
                log(f"  Visitas: {counter[0]}/{total} — parcial: {sum(visits.values()):,.0f}")

    # Procesar en batches de 20 threads simultáneos
    BATCH = 20
    for i in range(0, total, BATCH):
        batch = item_ids[i:i+BATCH]
        threads = [threading.Thread(target=fetch_one, args=(iid,)) for iid in batch]
        for t in threads: t.start()
        for t in threads: t.join()
        time.sleep(0.15)  # pausa entre batches para no saturar la API

    log(f"  Visitas: {total}/{total} — total: {sum(visits.values()):,.0f}")
    return visits

# ─── ÓRDENES ─────────────────────────────────────────────────────────────────

def get_orders_chunked(token, user_id):
    all_orders  = []
    chunk_start = DATE_FROM

    while chunk_start < DATE_TO:
        chunk_end = min(chunk_start + timedelta(days=1), DATE_TO)
        dfrom = chunk_start.strftime("%Y-%m-%dT%H:%M:%S.000-00:00")
        dto   = chunk_end.strftime("%Y-%m-%dT%H:%M:%S.000-00:00")

        probe = ml_get("https://api.mercadolibre.com/orders/search", token, params={
            "seller": user_id, "order.status": "paid",
            "order.date_created.from": dfrom, "order.date_created.to": dto,
            "sort": "date_asc", "offset": 0, "limit": 1,
        })
        if not probe:
            chunk_start = chunk_end; continue

        chunk_total = probe.get("paging", {}).get("total", 0)
        orders = []
        offset = 0
        while offset < chunk_total:
            data = ml_get("https://api.mercadolibre.com/orders/search", token, params={
                "seller": user_id, "order.status": "paid",
                "order.date_created.from": dfrom, "order.date_created.to": dto,
                "sort": "date_asc", "offset": offset, "limit": 50,
            })
            if not data: break
            batch = data.get("results", [])
            if not batch: break
            orders.extend(batch)
            if len(batch) < 50: break
            offset += 50
            time.sleep(0.3)

        all_orders.extend(orders)
        log(f"  {chunk_start.strftime('%d/%m')}: {len(orders)}/{chunk_total} órdenes")
        chunk_start = chunk_end
        time.sleep(0.3)

    return all_orders

def parse_orders(orders):
    units_by_item   = defaultdict(int)
    revenue_by_item = defaultdict(float)
    title_by_item   = {}
    price_by_item   = {}

    for order in orders:
        for item in order.get("order_items", []):
            iid   = str(item.get("item", {}).get("id", ""))
            qty   = item.get("quantity", 1)
            price = float(item.get("unit_price") or item.get("item", {}).get("price") or 0)
            title = item.get("item", {}).get("title", "")
            if iid:
                units_by_item[iid]   += qty
                revenue_by_item[iid] += qty * price
                if title:  title_by_item[iid]  = title
                if price:  price_by_item[iid]   = price

    return units_by_item, revenue_by_item, title_by_item, price_by_item

# ─── CATEGORIZACIÓN ───────────────────────────────────────────────────────────

def categorizar(visitas, ventas):
    # Criterios:
    # ⚫ Muerto:      <5 visitas y 0 ventas
    # 🔴 Invisible:  <15 visitas (y no cumple Muerto)
    # 🟡 Oportunidad: ≥15 visitas y conversión <1%
    # 🟢 Ganador:    ≥15 visitas y conversión ≥1%
    vis = visitas if isinstance(visitas, (int, float)) else 0
    ven = ventas  if isinstance(ventas,  (int, float)) else 0

    if vis < 5 and ven == 0:
        return ("⚫ Muerto",       f"0 ventas y {vis} visitas en {DAYS} días.", "Evaluar pausar, refactorizar o liquidar.", 4)
    elif vis < 15:
        return ("🔴 Invisible",    f"{ven} uds vendidas, {vis:,.0f} visitas. Muy baja visibilidad.", "Trabajar título, imágenes y precio para ganar tráfico.", 3)
    else:
        conv = ven / vis * 100
        if conv < 1:
            return ("🟡 Oportunidad", f"Buen tráfico ({vis:,.0f} visitas) pero solo {ven} uds vendidas. Conversión baja.", "Revisar precio, fotos y descripción vs competencia.", 2)
        else:
            return ("🟢 Ganador",     f"{ven} uds vendidas con {vis:,.0f} visitas. Conv: {conv:.2f}%.", "Mantener stock y escalar.", 1)

# ─── EXCEL ────────────────────────────────────────────────────────────────────

def write_excel(publications, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb   = Workbook()
    ws   = wb.active
    ws.title = "Semáforo ML Lavan"

    CLR_HEADER = "1F3864"
    CLR_WHITE  = "FFFFFF"
    CLR_TOTAL  = "E8F0FE"
    CAT_COLORS = {
        "🟢 Ganador":     ("CCFFCC","A5D6A7"),
        "🟡 Oportunidad": ("FFF9C4","FFF176"),
        "🔴 Invisible":   ("FFCDD2","EF9A9A"),
        "⚫ Muerto":      ("EEEEEE","E0E0E0"),
    }

    thin   = Side(style="thin",   color="CCCCCC")
    medium = Side(style="medium", color="888888")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_t  = Border(left=thin, right=thin, top=medium, bottom=thin)

    def fill(c):  return PatternFill("solid", start_color=c, fgColor=c)
    def font(bold=False, color="000000", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def al(h="left", v="center"): return Alignment(horizontal=h, vertical=v)

    total      = len(publications)
    ganadores  = sum(1 for p in publications if "Ganador"     in p["categoria"])
    opor       = sum(1 for p in publications if "Oportunidad" in p["categoria"])
    invisibles = sum(1 for p in publications if "Invisible"   in p["categoria"])
    muertos    = sum(1 for p in publications if "Muerto"      in p["categoria"])
    tot_vis    = sum(p["visitas"] for p in publications if isinstance(p.get("visitas"), (int,float)))
    tot_units  = sum(p["ventas"] for p in publications)
    tot_rev    = sum(p["ventas"] * p["price"] for p in publications)

    COLS = [
        ("Semáforo",          13), ("MLA",              18), ("SKU",          18),
        ("Título",            42), ("Estado",           10), ("¿Full?",        8),
        ("Stock total",       11), ("Precio ($)",       12), (f"Visitas {DAYS}d", 13),
        (f"Ventas {DAYS}d",   12), ("Conversión (%)",   13), ("Vel. sem.",    11),
        ("Revenue ($)",       14), ("Diagnóstico",      45), ("Acción",       40),
        ("URL ML",            40),
    ]
    NCOLS = len(COLS)
    last  = get_column_letter(NCOLS)

    # Fila 1: título
    ws.merge_cells(f"A1:{last}1")
    ws.cell(1,1, (
        f"🚦 SEMÁFORO ML CASA LAVAN — "
        f"Semana {last_monday.strftime('%d/%m')} al {last_sunday.strftime('%d/%m/%Y')}  |  "
        f"{total} publicaciones  |  "
        f"🟢 {ganadores}  🟡 {opor}  🔴 {invisibles}  ⚫ {muertos}"
    ))
    ws.cell(1,1).font = font(bold=True, color=CLR_WHITE, size=11)
    ws.cell(1,1).fill = fill(CLR_HEADER)
    ws.cell(1,1).alignment = al("center")
    ws.row_dimensions[1].height = 24

    # Fila 2: totales para validar
    ws.merge_cells(f"A2:{last}2")
    ws.cell(2,1, (
        f"TOTALES {DAYS}d — validar con ML:   "
        f"Visitas: {tot_vis:,.0f}   |   "
        f"Unidades: {tot_units:,.0f}   |   "
        f"Revenue: ${tot_rev:,.0f}"
    ))
    ws.cell(2,1).font = font(bold=True, color="1B5E20", size=10)
    ws.cell(2,1).fill = fill("E8F5E9")
    ws.cell(2,1).alignment = al("center")
    ws.row_dimensions[2].height = 18

    # Fila 3: criterios
    ws.merge_cells(f"A3:{last}3")
    ws.cell(3,1, (
        "Criterios:  ⚫ Muerto: <5 visitas y 0 ventas  ·  "
        "🔴 Invisible: <15 visitas  ·  "
        "🟡 Oportunidad: ≥15 visitas y conversión <1%  ·  "
        "🟢 Ganador: ≥15 visitas y conversión ≥1%"
    ))
    ws.cell(3,1).font = font(color="555555", size=9)
    ws.cell(3,1).fill = fill("F5F5F5")
    ws.cell(3,1).alignment = al("center")
    ws.row_dimensions[3].height = 14

    # Fila 4: headers
    for ci,(label,width) in enumerate(COLS,1):
        c = ws.cell(4, ci, label)
        c.font = font(bold=True, color=CLR_WHITE)
        c.fill = fill(CLR_HEADER)
        c.alignment = al("center")
        c.border = brd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[4].height = 18
    ws.freeze_panes = "A5"

    # Ordenar: Oportunidad > Invisible > Ganador > Muerto
    orden = {"🟡 Oportunidad":1,"🔴 Invisible":2,"🟢 Ganador":3,"⚫ Muerto":4}
    publications.sort(key=lambda p: (
        orden.get(p["categoria"], 9),
        -(p["visitas"] if isinstance(p.get("visitas"),(int,float)) else 0)
    ))

    for ri, pub in enumerate(publications, 5):
        cat    = pub["categoria"]
        bg_n, bg_a = CAT_COLORS.get(cat, ("EEEEEE","E0E0E0"))
        bg     = bg_a if ri%2==0 else bg_n
        vis    = pub.get("visitas")
        ventas = pub.get("ventas", 0)
        conv   = round(ventas/vis*100, 2) if vis and vis > 0 else None
        vel    = round(ventas/DAYS*7, 2)
        rev    = round(ventas * pub.get("price", 0))

        row_data = [
            cat, pub["item_id"], pub.get("sku",""), pub.get("title",""),
            pub.get("status",""), pub.get("logistica",""),
            pub.get("stock",0), pub.get("price",0),
            vis if vis is not None else "—",
            ventas,
            conv if conv is not None else "—",
            vel, rev,
            pub.get("diagnostico",""), pub.get("accion",""), pub.get("permalink",""),
        ]

        for ci, val in enumerate(row_data, 1):
            key = COLS[ci-1][0]
            c   = ws.cell(ri, ci)
            c.fill = fill(bg); c.border = brd

            if key == "URL ML":
                if val and str(val).startswith("http"):
                    c.value = val; c.hyperlink = val
                    c.font  = Font(name="Arial", size=9, color="1155CC", underline="single")
                else:
                    c.value = "—"; c.font = font(color="AAAAAA", size=9)
                c.alignment = al("left")
            elif key == "Semáforo":
                c.value = val; c.font = font(bold=True, size=10); c.alignment = al("center")
            elif key == "¿Full?":
                c.value = val; c.font = font(size=10); c.alignment = al("center")
                c.value = val; c.alignment = al("center")
                c.font = font(bold=True, color="1B5E20", size=9) if val=="active" else font(color="B71C1C", size=9)
            elif key == "Stock total":
                c.value = val; c.number_format = '#,##0'; c.alignment = al("center")
                c.font = font(bold=True, color="B71C1C", size=9) if val==0 else font(size=9)
            elif key in ("Precio ($)","Revenue ($)"):
                c.value = val; c.font = font(size=9)
                c.number_format = '#,##0'; c.alignment = al("right")
            elif key == "Conversión (%)":
                c.value = val; c.font = font(size=9)
                if isinstance(val,(int,float)): c.number_format = '0.00"%"'
                c.alignment = al("center")
            elif key in (f"Visitas {DAYS}d", f"Ventas {DAYS}d", "Vel. sem."):
                c.value = val; c.font = font(size=9)
                if isinstance(val,(int,float)):
                    c.number_format = '0.0' if "Vel" in key else '#,##0'
                c.alignment = al("center")
            elif key in ("Diagnóstico","Acción","Título"):
                c.value = val; c.font = font(size=9); c.alignment = al("left")
            else:
                c.value = val; c.font = font(size=9); c.alignment = al("left")

        ws.row_dimensions[ri].height = 14

    # Fila total
    ri_t = 5 + len(publications)
    ws.cell(ri_t,1,"TOTAL").font = font(bold=True)
    ws.cell(ri_t,9, tot_vis).number_format   = '#,##0'
    ws.cell(ri_t,10,tot_units).number_format = '#,##0'
    ws.cell(ri_t,13,tot_rev).number_format   = '#,##0'
    for ci in range(1, NCOLS+1):
        c = ws.cell(ri_t, ci)
        c.fill = fill(CLR_TOTAL); c.border = brd_t; c.font = font(bold=True, size=9)
        if ci in (9,10,13): c.alignment = al("right")
    ws.row_dimensions[ri_t].height = 16

    ws.auto_filter.ref = f"A4:{last}{ri_t-1}"
    wb.save(out_path)

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'═'*56}")
    print(f"  SEMÁFORO ML CASA LAVAN")
    print(f"  Semana: {last_monday.strftime('%d/%m/%Y')} (lun) → {last_sunday.strftime('%d/%m/%Y')} (dom)")
    print(f"  Config: {config_path}")
    print(f"{'═'*56}\n")

    user_id = cfg_lavan["user_id"]

    log("Renovando token ML Lavan...")
    token = ml_refresh(cfg_lavan)
    log("✓ Token listo")

    # 1. Órdenes
    log(f"Trayendo órdenes de los últimos {DAYS} días...")
    orders = get_orders_chunked(token, user_id)
    units_by_item, revenue_by_item, title_by_item, price_by_item = parse_orders(orders)
    tot_units_ord = sum(units_by_item.values())
    tot_rev_ord   = sum(revenue_by_item.values())
    log(f"✓ {len(orders)} órdenes | {tot_units_ord:,.0f} unidades | ${tot_rev_ord:,.0f} revenue")

    # 2. Items activos — scroll sin límite
    log("Trayendo IDs de publicaciones activas (scroll)...")
    item_ids = get_all_item_ids(token, user_id)
    log(f"✓ {len(item_ids)} publicaciones activas")

    # 3. Detalle
    log("Trayendo detalle de items...")
    details = get_items_detail(token, item_ids)
    log(f"✓ Detalle de {len(details)} items")

    # 4. Visitas — batch con fecha simple
    log(f"Trayendo visitas ({DATE_FROM_SIMPLE} → {DATE_TO_SIMPLE})...")
    visits = get_visits(token, item_ids)
    tot_vis = sum(v for v in visits.values() if isinstance(v,(int,float)))
    log(f"✓ Visitas totales: {tot_vis:,.0f}")

    # 5. Construir publicaciones
    log("Construyendo análisis...")
    def logistic_label(lt):
        return "✅" if lt == "fulfillment" else "—"

    publications = []
    for item_id in item_ids:
        det    = details.get(item_id, {})
        vis    = visits.get(item_id)
        ventas = units_by_item.get(item_id, 0)
        price  = det.get("price") or price_by_item.get(item_id, 0)
        title  = det.get("title") or title_by_item.get(item_id, "")
        cat, diag, accion, _ = categorizar(vis, ventas)

        publications.append({
            "item_id":    item_id,
            "sku":        det.get("sku",""),
            "title":      title,
            "price":      price,
            "permalink":  det.get("permalink",""),
            "status":     det.get("status",""),
            "stock":      det.get("stock",0),
            "logistica":  logistic_label(det.get("logistic_type","")),
            "visitas":    vis,
            "ventas":     ventas,
            "categoria":  cat,
            "diagnostico": diag,
            "accion":     accion,
        })

    # 6. Excel
    fecha    = f"{last_monday.strftime('%Y-%m-%d')}_al_{last_sunday.strftime('%Y-%m-%d')}"
    out_dir  = config_path.parent / "semaforo"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"semaforo_ml_lavan_{fecha}.xlsx"
    log("Generando Excel...")
    write_excel(publications, out_path)

    ganadores  = sum(1 for p in publications if "Ganador"     in p["categoria"])
    opor       = sum(1 for p in publications if "Oportunidad" in p["categoria"])
    invisibles = sum(1 for p in publications if "Invisible"   in p["categoria"])
    muertos    = sum(1 for p in publications if "Muerto"      in p["categoria"])

    print(f"\n  {'━'*50}")
    print(f"  ✅ {out_path.name}")
    print(f"  {'━'*50}")
    print(f"  TOTALES — comparar con ML dashboard:")
    print(f"    Publicaciones:     {len(publications):,.0f}")
    print(f"    Visitas {DAYS}d:       {tot_vis:,.0f}")
    print(f"    Unidades {DAYS}d:      {tot_units_ord:,.0f}")
    print(f"    Revenue {DAYS}d:       ${tot_rev_ord:,.0f}")
    print(f"  {'━'*50}")
    print(f"  🟢 {ganadores} Ganadoras  🟡 {opor} Oportunidades  🔴 {invisibles} Invisibles  ⚫ {muertos} Muertas")
    print(f"\n{'═'*56}\n")

if __name__ == "__main__":
    main()
